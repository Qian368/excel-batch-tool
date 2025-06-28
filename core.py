#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel处理核心模块
提供Excel文件的各种批量处理功能
"""

import os
import traceback
from pathlib import Path
import shutil
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter


class ExcelProcessor:
    """
    Excel处理核心类，提供各种Excel批量处理功能
    """
    
    def __init__(self):
        self.workbooks = {}
        self.output_dir = None
        self.temp_files = {}
        self.selected_worksheets = {}  # 用户选择的工作表
        
        # 清理可能存在的临时文件
        self._cleanup_temp_files()
    
    def set_output_dir(self, output_dir):
        """
        设置Excel文件的输出目录
        
        Args:
            output_dir: 输出目录路径
        """
        if not output_dir:
            raise ValueError("输出目录不能为空")
            
        output_path = Path(output_dir)
        if not output_path.exists():
            output_path.mkdir(parents=True)
            
        self.output_dir = str(output_path)
    
    def set_selected_worksheets(self, selected_worksheets):
        """
        设置用户选择的工作表
        
        Args:
            selected_worksheets: 文件路径到工作表名称列表的映射
        """
        self.selected_worksheets = selected_worksheets
    
    def load_workbooks(self, file_paths):
        """
        加载Excel工作簿，并复制到临时工作区
        
        Args:
            file_paths: Excel文件路径列表
        """
        for file_path in file_paths:
            try:
                # 创建临时文件副本
                temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                shutil.copy2(file_path, temp_path)
                self.temp_files[file_path] = temp_path
                
                # 加载临时文件的工作簿
                wb = openpyxl.load_workbook(temp_path, data_only=False)
                self.workbooks[file_path] = wb
            except Exception as e:
                error_msg = f"加载文件 {file_path} 失败: {str(e)}"
                print(error_msg)
                raise ValueError(error_msg)
    
    def save_workbooks(self):
        """
        保存所有工作簿到指定的输出目录
        """
        if not self.output_dir:
            raise ValueError("未指定输出目录")
            
        # 确保输出目录存在且可写
        output_path = Path(self.output_dir)
        if not output_path.exists():
            try:
                output_path.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                raise PermissionError(f"无法创建输出目录 {self.output_dir}: {str(e)}")
                
        if not os.access(self.output_dir, os.W_OK):
            raise PermissionError(f"输出目录 {self.output_dir} 不可写")
            
        for file_path, wb in self.workbooks.items():
            try:
                # 获取文件名并构建新的保存路径
                file_name = Path(file_path).name
                output_path = Path(self.output_dir) / file_name
                
                # 检查目标文件是否已存在且可写
                if output_path.exists() and not os.access(str(output_path), os.W_OK):
                    raise PermissionError(f"目标文件 {output_path} 已存在且不可写")
                    
                # 保存工作簿
                wb.save(output_path)
                
            except PermissionError as e:
                print(f"权限错误: {str(e)}")
                raise
            except Exception as e:
                print(f"保存文件 {file_path} 失败: {str(e)}")
                raise
            finally:
                # 清理临时文件
                self._cleanup_temp_file(file_path)
    
    def convert_formulas_to_values(self, file_paths):
        """
        将公式转换为值
        
        Args:
            file_paths: 要处理的Excel文件路径列表
        """
        for file_path in file_paths:
            try:
                # 确保使用工作簿副本，而不是直接修改原始工作簿
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 先用data_only=True加载以计算公式
                    wb_data = openpyxl.load_workbook(temp_path, data_only=True)
                    
                    # 再用data_only=False加载以获取原始内容
                    wb_formula = openpyxl.load_workbook(temp_path, data_only=False)
                else:
                    # 如果已经有工作簿副本，则使用它
                    # 获取临时文件路径
                    temp_path = self.temp_files.get(file_path)
                    if not temp_path:
                        # 如果没有临时文件，创建一个
                        temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                        shutil.copy2(file_path, temp_path)
                        self.temp_files[file_path] = temp_path
                    
                    # 先用data_only=True加载以计算公式
                    wb_data = openpyxl.load_workbook(temp_path, data_only=True)
                    
                    # 使用已有的工作簿
                    wb_formula = wb
                
                # 确定要处理的工作表
                sheets_to_process = []
                if file_path in self.selected_worksheets and self.selected_worksheets[file_path]:
                    # 如果用户选择了特定的工作表，则只处理这些工作表
                    for sheet_name in self.selected_worksheets[file_path]:
                        if sheet_name in wb_formula.sheetnames:
                            sheets_to_process.append(sheet_name)
                else:
                    # 否则处理所有工作表
                    sheets_to_process = wb_formula.sheetnames
                
                for sheet_name in sheets_to_process:
                    sheet_formula = wb_formula[sheet_name]
                    sheet_data = wb_data[sheet_name]
                    
                    for row in range(1, sheet_formula.max_row + 1):
                        for col in range(1, sheet_formula.max_column + 1):
                            cell_formula = sheet_formula.cell(row=row, column=col)
                            cell_data = sheet_data.cell(row=row, column=col)
                            
                            if cell_formula.data_type == 'f':  # 如果是公式
                                cell_formula.value = cell_data.value
                
                # 保存修改后的工作簿到临时文件
                if temp_path:
                    wb_formula.save(temp_path)
                self.workbooks[file_path] = wb_formula
                
            except Exception as e:
                print(f"处理文件 {file_path} 失败: {str(e)}")
                raise
    
    def _cleanup_temp_files(self):
        """
        清理所有临时文件
        """
        for file_path, temp_path in list(self.temp_files.items()):
            try:
                if Path(temp_path).exists():
                    Path(temp_path).unlink()
                del self.temp_files[file_path]
            except Exception as e:
                print(f"清理临时文件 {temp_path} 失败: {str(e)}")
    
    def _cleanup_temp_file(self, file_path):
        """
        清理指定文件的临时文件
        
        Args:
            file_path: 原始文件路径
        """
        if file_path in self.temp_files:
            try:
                temp_path = self.temp_files[file_path]
                if Path(temp_path).exists():
                    Path(temp_path).unlink()
                del self.temp_files[file_path]
            except Exception as e:
                print(f"清理临时文件 {self.temp_files[file_path]} 失败: {str(e)}")
    
    def change_font_color(self, file_paths, color, range_mode='specific', range_str=None):
        """
        修改单元格字体颜色
        
        Args:
            file_paths: 要处理的Excel文件路径列表
            color: 字体颜色名称
            range_mode: 范围模式，'specific'表示指定范围，'entire_sheet'表示整个工作表
            range_str: 当range_mode为'specific'时，指定要处理的单元格范围
            
        Returns:
            dict: 包含操作结果的字典
        """
        from cell_format_module import change_font_color
        return change_font_color(self, file_paths, color, range_mode, range_str)
    
    def change_fill_color(self, file_paths, color, range_mode='specific', range_str=None):
        """
        修改单元格填充颜色
        
        Args:
            file_paths: 要处理的Excel文件路径列表
            color: 填充颜色名称
            range_mode: 范围模式，'specific'表示指定范围，'entire_sheet'表示整个工作表
            range_str: 当range_mode为'specific'时，指定要处理的单元格范围
            
        Returns:
            dict: 包含操作结果的字典
        """
        from cell_format_module import change_fill_color
        return change_fill_color(self, file_paths, color, range_mode, range_str)
    
    def add_border(self, file_paths, range_str, border_style='thin'):
        """
        添加单元格边框
        
        Args:
            file_paths: 要处理的Excel文件路径列表
            range_str: 指定要处理的单元格范围
            border_style: 边框样式，默认为'thin'
            
        Returns:
            dict: 包含操作结果的字典
        """
        from cell_format_module import add_border
        return add_border(self, file_paths, range_str, border_style)
    
    def remove_border(self, file_paths, range_str):
        """
        移除单元格边框
        
        Args:
            file_paths: 要处理的Excel文件路径列表
            range_str: 指定要处理的单元格范围
            
        Returns:
            dict: 包含操作结果的字典
        """
        from cell_format_module import remove_border
        return remove_border(self, file_paths, range_str)
    
    def modify_cell_content(self, file_paths, cell_position, new_content):
        """
        修改单元格内容
        
        Args:
            file_paths: 要处理的Excel文件路径列表
            cell_position: 单元格位置，如'A1'或'A1:B2'或'A1,C3:D4'
            new_content: 新的单元格内容
            
        Returns:
            dict: 包含操作结果的字典
        """
        try:
            from cell_utils import parse_cell_range, process_cell_ranges, get_cell_coordinates
            
            # 定义修改单元格内容的函数
            def set_cell_content(cell, content=new_content):
                cell.value = content
            
            for file_path in file_paths:
                # 确保使用工作簿副本
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                # 确定要处理的工作表
                sheets_to_process = []
                if file_path in self.selected_worksheets and self.selected_worksheets[file_path]:
                    # 如果用户选择了特定的工作表，则只处理这些工作表
                    for sheet_name in self.selected_worksheets[file_path]:
                        if sheet_name in wb.sheetnames:
                            sheets_to_process.append(sheet_name)
                else:
                    # 否则处理所有工作表
                    sheets_to_process = wb.sheetnames
                
                for sheet_name in sheets_to_process:
                    sheet = wb[sheet_name]
                    
                    if not cell_position:
                        raise ValueError("未指定单元格位置")
                    
                    # 替换中文符号
                    cell_position = cell_position.replace('，', ',').replace('：', ':')
                    
                    # 解析单元格范围
                    cell_ranges = parse_cell_range(cell_position)
                    
                    # 特殊处理合并单元格的情况
                    for start_cell, end_cell in cell_ranges:
                        # 检查是否是合并单元格
                        if start_cell == end_cell:
                            # 单个单元格，检查是否在合并单元格内
                            row, col = get_cell_coordinates(start_cell)
                            for merged_range in sheet.merged_cells.ranges:
                                if (merged_range.min_row <= row <= merged_range.max_row and 
                                    merged_range.min_col <= col <= merged_range.max_col):
                                    # 如果在合并单元格内，只修改左上角单元格的值
                                    sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value = new_content
                                    break
                    
                    # 使用通用方法处理单元格范围，包括合并单元格
                    result = process_cell_ranges(sheet, cell_ranges, set_cell_content)
                    if not result['success']:
                        raise ValueError(f"修改单元格内容失败: {result['message']}")
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
            
            return {
                'success': True,
                'message': f"修改单元格内容（位置：{cell_position}）执行成功"
            }
        except Exception as e:
            return {
                'success': False,
                'message': f"修改单元格内容执行失败: {str(e)}"
            }

    
    def _get_data_range(self, sheet):
        """
        获取工作表中有数据的范围
        
        Args:
            sheet: 工作表对象
            
        Returns:
            tuple: (min_row, max_row, min_col, max_col)
        """
        min_row = min_col = float('inf')
        max_row = max_col = 0
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    min_row = min(min_row, row)
                    max_row = max(row, max_row)
                    min_col = min(min_col, col)
                    max_col = max(col, max_col)
        
        return (min_row if min_row != float('inf') else 1,
                max_row if max_row != 0 else sheet.max_row,
                min_col if min_col != float('inf') else 1,
                max_col if max_col != 0 else sheet.max_column)
    
    def get_intersection(self, range1, range2):
        """
        计算两个单元格范围的交集
        
        Args:
            range1: 元组 (min_row1, min_col1, max_row1, max_col1)
            range2: 元组 (min_row2, min_col2, max_row2, max_col2)
        
        Returns:
            tuple: 交集区域的范围 (min_row, min_col, max_row, max_col) 或 None（如果没有交集）
        """
        min_row1, min_col1, max_row1, max_col1 = range1
        min_row2, min_col2, max_row2, max_col2 = range2
        
        # 如果一个范围在另一个范围的完全左边或右边，则没有交集
        if max_col1 < min_col2 or min_col1 > max_col2:
            return None
        
        # 如果一个范围在另一个范围的完全上边或下边，则没有交集
        if max_row1 < min_row2 or min_row1 > max_row2:
            return None
        
        # 计算交集区域
        intersection_min_row = max(min_row1, min_row2)
        intersection_min_col = max(min_col1, min_col2)
        intersection_max_row = min(max_row1, max_row2)
        intersection_max_col = min(max_col1, max_col2)
        
        return (intersection_min_row, intersection_min_col, intersection_max_row, intersection_max_col)
    
    def get_range_from_reference(self, range_ref):
        """
        从单元格引用获取范围的行列数值
        
        Args:
            range_ref: 字符串，如'A1:B2'或'A1'
            
        Returns:
            tuple: (min_row, min_col, max_row, max_col)
        """
        from openpyxl.utils import range_boundaries
        
        try:
            # 使用openpyxl的函数解析单元格范围
            min_col, min_row, max_col, max_row = range_boundaries(range_ref)
            return (min_row, min_col, max_row, max_col)
        except Exception:
            return None
    
    def find_intersections(self, ws, target_range):
        """
        查找与目标范围相交的合并单元格
        
        Args:
            ws: openpyxl工作表对象
            target_range: 目标单元格范围的引用字符串
        
        Returns:
            list: 相交的合并单元格列表，每个元素为(合并单元格范围, 交集范围)元组
        """
        target = self.get_range_from_reference(target_range)
        
        if not target:
            return []
            
        # 查找与目标范围有交集的合并单元格
        intersections = []
        for merged_range in ws.merged_cells.ranges:
            current_range = (
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col
            )
            
            intersection = self.get_intersection(target, current_range)
            if intersection:
                # 将交集区域转换为Excel引用格式
                from openpyxl.utils import get_column_letter
                intersection_start = f"{get_column_letter(intersection[1])}{intersection[0]}"
                intersection_end = f"{get_column_letter(intersection[3])}{intersection[2]}"
                intersection_range = intersection_start if intersection_start == intersection_end else f"{intersection_start}:{intersection_end}"
                
                intersections.append((merged_range, intersection_range))
        
        return intersections
    
    def _get_merged_cells_in_range(self, sheet, start_row, end_row, start_col, end_col):
        """
        获取指定范围内的合并单元格
        
        Args:
            sheet: 工作表对象
            start_row: 起始行
            end_row: 结束行
            start_col: 起始列
            end_col: 结束列
            
        Returns:
            list: 合并单元格范围列表
        """
        merged_ranges = []
        for merged_range in sheet.merged_cells.ranges:
            # 检查合并单元格是否与指定范围有交集
            if not (merged_range.max_row < start_row or 
                    merged_range.min_row > end_row or 
                    merged_range.max_col < start_col or 
                    merged_range.min_col > end_col):
                merged_ranges.append(merged_range)
        return merged_ranges
    
    def process_merged_cells(self, file_paths, action='unmerge', mode='all', range_str=None):
        """
        处理合并单元格（合并或拆分）
        
        Args:
            file_paths: 要处理的Excel文件路径列表
            action: 操作类型，'merge'表示合并，'unmerge'表示拆分，'keep_value'表示保留值但拆分
            mode: 处理模式，'all'表示处理整个工作表，'specific'表示处理指定范围
            range_str: 当mode为'specific'时，指定要处理的单元格范围
        """
        for file_path in file_paths:
            try:
                # 确保使用工作簿副本，而不是直接修改原始工作簿
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    if mode == 'specific' and range_str:
                        # 处理指定范围的合并单元格
                        # 解析单元格范围
                        try:
                            # 处理单个单元格的情况
                            if ':' not in range_str:
                                # 解析单个单元格坐标
                                cell_coord = openpyxl.utils.cell.coordinate_to_tuple(range_str)
                                min_row, min_col = cell_coord
                                max_row, max_col = cell_coord
                            else:
                                # 处理范围的情况
                                start_cell, end_cell = range_str.split(':')
                                start_coord = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
                                end_coord = openpyxl.utils.cell.coordinate_to_tuple(end_cell)
                                
                                min_row, min_col = start_coord
                                max_row, max_col = end_coord
                            
                            # 找出所有与指定范围有交集的合并单元格
                            overlapping_ranges = []
                            for merged_range in list(sheet.merged_cells.ranges):
                                merged_min_row, merged_min_col = merged_range.min_row, merged_range.min_col
                                merged_max_row, merged_max_col = merged_range.max_row, merged_range.max_col
                                
                                # 检查是否有重叠
                                if not (merged_max_row < min_row or merged_min_row > max_row or 
                                        merged_max_col < min_col or merged_min_col > max_col):
                                    overlapping_ranges.append(merged_range)
                            
                            # 如果没有找到任何重叠的合并单元格，则报告错误
                            # 但是不要立即抛出异常，而是返回一个标志，让调用者决定如何处理
                            if not overlapping_ranges:
                                # 这里不再抛出异常，而是返回一个空列表，表示没有找到重叠的合并单元格
                                # 这样processing.py中的代码可以继续处理其他文件
                                continue
                            
                            # 拆分所有重叠的合并单元格
                            for merged_range in overlapping_ranges:
                                # 如果需要保留值，先获取左上角单元格的值
                                top_left_value = None
                                if action == 'keep_value':
                                    top_left_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                                
                                # 拆分合并单元格
                                sheet.unmerge_cells(str(merged_range))
                                
                                # 如果需要保留值，填充到所有单元格
                                if action == 'keep_value' and top_left_value is not None:
                                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                                            sheet.cell(row=row, column=col).value = top_left_value
                                
                        except ValueError as ve:
                            # 传递自定义的ValueError
                            raise ve
                        except Exception as e:
                            # 如果解析单元格范围出错，抛出更具体的错误
                            raise ValueError(f"无效的单元格范围格式: {range_str}, 错误: {str(e)}")
                        
                        # 注意：合并单元格的值保留逻辑已经在拆分合并单元格的循环中处理了
                    else:
                        # 处理整个工作表的所有合并单元格
                        merged_ranges = list(sheet.merged_cells.ranges)
                        
                        for merged_range in merged_ranges:
                            # 获取左上角单元格的值
                            top_left_value = sheet.cell(
                                row=merged_range.min_row, 
                                column=merged_range.min_col
                            ).value
                            
                            # 拆分合并单元格
                            sheet.unmerge_cells(str(merged_range))
                            
                            if action == 'keep_value':
                                # 填充到所有单元格
                                for row in range(merged_range.min_row, merged_range.max_row + 1):
                                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                                        sheet.cell(row=row, column=col).value = top_left_value
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"处理文件 {file_path} 失败: {str(e)}")
                raise
    
    def _process_merged_cells_in_range(self, sheet, merged_ranges, merge_mode='ignore'):
        """
        根据不同模式处理指定范围内的合并单元格
        
        Args:
            sheet: 工作表对象
            merged_ranges: 合并单元格范围列表
            merge_mode: 合并单元格处理模式，可选值：
                       - ignore: 忽略合并单元格，直接执行操作
                       - unmerge_only: 仅拆分合并单元格，不保留值
                       - unmerge_keep_value: 拆分合并单元格并保留值
        """
        if merge_mode == 'ignore':
            return

        #遍历所有合并单元格    
        for merged_range in merged_ranges:
            # 获取左上角单元格的值
            top_left_value = None
            #if merge_mode == 'unmerge_keep_value':
            top_left_value = sheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            ).value
            
            # 如果是unmerge_keep_value模式，则填充值到所有单元格
            if merge_mode == 'unmerge_keep_value':
                # 拆分合并单元格
                sheet.unmerge_cells(str(merged_range))
                # 将值填充到所有相关单元格
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        sheet.cell(row=row, column=col).value = top_left_value
            # 如果是unmerge_only模式，则清空所有单元格的值
            elif merge_mode == 'unmerge_only':
                # 拆分合并单元格
                sheet.unmerge_cells(str(merged_range))
                # 将值保留在左上角单元格
                sheet.cell(merged_range.min_row, merged_range.min_col).value = top_left_value
    
    def insert_rows(self, file_paths, sheet_indexes, position, count=1):
        """
        插入行
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            position: 插入位置（行号）
            count: 插入行数
        """
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                for sheet_index in sheet_indexes:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        sheet.insert_rows(position, count)
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中插入行失败: {str(e)}")
                raise
    
    def delete_rows(self, file_paths, sheet_indexes, position, count=1, merge_mode='ignore'):
        """
        删除行
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            position: 删除位置（行号）
            count: 删除行数
            merge_mode: 合并单元格处理模式
        """
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                # 确定要处理的工作表
                sheets_to_process = []
                if file_path in self.selected_worksheets and self.selected_worksheets[file_path]:
                    # 如果用户选择了特定的工作表，则只处理这些工作表
                    for sheet_name in self.selected_worksheets[file_path]:
                        if sheet_name in wb.sheetnames:
                            sheets_to_process.append(wb.sheetnames.index(sheet_name))
                else:
                    # 否则使用传入的sheet_indexes
                    sheets_to_process = sheet_indexes
                
                for sheet_index in sheets_to_process:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        
                        # 获取数据范围
                        data_range = self._get_data_range(sheet)
                        
                        try:
                            # 获取要删除的行范围内的合并单元格
                            merged_ranges_intersect = self._get_merged_cells_in_range(
                                sheet, position, position + count - 1,
                                data_range[2], data_range[3]  # 列范围
                            )
                            
                            # 获取下方的合并单元格
                            merged_ranges_below = []
                            for merged_range in sheet.merged_cells:
                                if merged_range.min_row > position + count - 1:
                                    merged_ranges_below.append(merged_range)

                            # 合并交集和要删除的行下方区域的合并单元格
                            merged_ranges = list(set(merged_ranges_intersect + merged_ranges_below))

                            # 记录不同merge_mode模式下要处理的合并单元格信息
                            merge_info_list = []
                            
                            # 不是ignore要处理下方合并单元格
                            if merged_ranges and merge_mode != 'ignore':
                                self._process_merged_cells_in_range(sheet, merged_ranges, merge_mode)
                                
                                if merged_ranges_below:
                                    print("下方合并单元格信息:")
                                    # 记录下方的合并单元格的信息（删除行前记录）
                                    for rng in merged_ranges_below:
                                        top_left = sheet.cell(rng.min_row, rng.min_col)
                                        value = top_left.value if top_left else None
                                        print(f"  合并区域: {rng.coord}, 值: '{value}'") 

                                        # 记录合并信息
                                        merge_info_list.append({
                                            'min_row': rng.min_row,
                                            'max_row': rng.max_row,
                                            'min_col': rng.min_col,
                                            'max_col': rng.max_col,
                                            'value': value,
                                            'range_str': str(rng)
                                        })                         

                            # 模式为"ignore"时，要处理交集和删除行下方的单元格
                            elif merged_ranges and merge_mode == 'ignore':    
                                print(f"检测到 {len(merged_ranges)} 个受影响的(包括交集和下方)合并单元格")
                                
                                # 记录所有合并区域的信息（删除行前记录）
                                for rng in merged_ranges:
                                    top_left = sheet.cell(rng.min_row, rng.min_col)
                                    value = top_left.value if top_left else None
                                    print(f"  合并区域: {rng.coord}, 值: '{value}'")
                                    
                                    # 记录合并信息
                                    merge_info_list.append({
                                        'min_row': rng.min_row,
                                        'max_row': rng.max_row,
                                        'min_col': rng.min_col,
                                        'max_col': rng.max_col,
                                        'value': value,
                                        'range_str': str(rng)
                                    })

                            # 根据已经已载入不同值的merge_info_list重建合并单元格
                            if merge_info_list:    
                                # 1. 先解除所有合并
                                print("第一步: 解除所有合并")
                                for info in merge_info_list:
                                    try:
                                        if info['range_str'] in [str(m) for m in sheet.merged_cells]:
                                            sheet.unmerge_cells(info['range_str'])
                                            print(f"  ✓ 已解除: {info['range_str']}")
                                    except Exception as e:
                                        print(f"  ⚠ 解除警告: {str(e)}")
                                
                                # 2. 执行删除行操作（安全删除——从下往上删除）
                                print("\n第二步: 删除行")
                                rows_to_delete = list(range(position, position + count))
                                for row_idx in sorted(rows_to_delete, reverse=True):
                                    print(f"  删除行: {row_idx}")
                                    sheet.delete_rows(row_idx, 1)
                                
                                # 3. 重建所有合并单元格
                                print("\n第三步: 重建合并区域")
                                for info in merge_info_list:
                                    # 计算新的行范围
                                    new_min_row = info['min_row'] - len([r for r in rows_to_delete if r < info['min_row']])
                                    new_max_row = info['max_row'] - len([r for r in rows_to_delete if r <= info['max_row']]) 
                                    
                                    if any(r == info['min_row'] for r in rows_to_delete):
                                        print(f"  × 合并区域上边界为删除行，无需合并和恢复值")
                                        continue                                    
                                    # 检查新位置是否有效
                                    if new_min_row < 1 or new_max_row < 1 or new_min_row > new_max_row:
                                        print(f"  × 跳过无效范围: 原范围为：MIN：{info['min_row']} ～ MAX：{info['max_row']} => 新范围为：MIN：{new_min_row} ～ MAX：{new_max_row}")
                                        continue

                                    if  new_min_row == new_max_row and info['min_col'] == info['max_col']:
                                        print(f"  ✓ 删除后只有一个单元格，无需合并，原来的值为：{info['value']}")
                                        continue 
                                                                      
                                    # 有效合并范围执行合并操作
                                    if new_min_row >= 1 and new_max_row >= 1 and new_min_row < new_max_row:                                    
                                        # 创建新的合并范围
                                        new_range_str = (
                                            f"{get_column_letter(info['min_col'])}{new_min_row}:"
                                            f"{get_column_letter(info['max_col'])}{new_max_row}"
                                        )
                                        
                                        try:
                                            # 创建新的合并区域
                                            sheet.merge_cells(new_range_str)
                                            print(f"  ✓ 创建新合并: {new_range_str}")

                                            # 设置左上角单元格的值
                                            sheet.cell(new_min_row, info['min_col']).value = info['value']                                            
                                            
                                            print(f"  ✓ 值已恢复: '{info['value']}'")
                                        
                                        except Exception as e:
                                            print(f"  ⚠ 创建合并失败: {str(e)}")
                                
                                print("合并区域重建完成") 

                            # 如果都没有交集或者下方合并单元格需要处理，则直接删除行
                            else:
                                # 直接执行————删除行操作（安全删除——从下往上删除）
                                print("\n直接删除行：")
                                rows_to_delete = list(range(position, position + count))
                                for row_idx in sorted(rows_to_delete, reverse=True):
                                    print(f"  删除行: {row_idx}")
                                    sheet.delete_rows(row_idx, 1)   

                        except ValueError as ve:
                            # 提供更具体的错误信息
                            raise ValueError(f"删除行失败: {str(ve)}")
                        except Exception as e:
                            # 捕获其他可能的错误
                            raise ValueError(f"删除行时发生错误: {str(e)}")
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中删除行失败: {str(e)}")
                raise
    
    def insert_columns(self, file_paths, sheet_indexes, position, count=1):
        """
        插入列
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            position: 插入位置（列号）
            count: 插入列数
        """
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                for sheet_index in sheet_indexes:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        sheet.insert_cols(position, count)
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中插入列失败: {str(e)}")
                raise
    
    def delete_columns(self, file_paths, sheet_indexes, position, count=1, merge_mode='ignore'):
        """
        删除列
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            position: 删除位置（列号）
            count: 删除列数
            merge_mode: 合并单元格处理模式
        """
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                # 确定要处理的工作表
                sheets_to_process = []
                if file_path in self.selected_worksheets and self.selected_worksheets[file_path]:
                    # 如果用户选择了特定的工作表，则只处理这些工作表
                    for sheet_name in self.selected_worksheets[file_path]:
                        if sheet_name in wb.sheetnames:
                            sheets_to_process.append(wb.sheetnames.index(sheet_name))
                else:
                    # 否则使用传入的sheet_indexes
                    sheets_to_process = sheet_indexes
                
                for sheet_index in sheets_to_process:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        
                        # 获取数据范围
                        data_range = self._get_data_range(sheet)
                        
                        try:
                            # 获取要删除的列范围内的合并单元格
                            merged_ranges_intersect = self._get_merged_cells_in_range(
                                sheet, data_range[0], data_range[1],
                                position, position + count - 1
                            )
                            
                            # 获取右侧的合并单元格
                            merged_ranges_right = []
                            for merged_range in sheet.merged_cells:
                                if merged_range.min_col > position + count - 1:
                                    merged_ranges_right.append(merged_range)

                            # 合并交集和要删除的列的右侧区域的合并单元格
                            merged_ranges = list(set(merged_ranges_intersect + merged_ranges_right))

                            # 记录不同merge_mode模式下要处理的合并单元格信息
                            merge_info_list = []
                            
                            # 不是ignore要处理右侧合并单元格
                            if merged_ranges and merge_mode != 'ignore':
                                self._process_merged_cells_in_range(sheet, merged_ranges, merge_mode)
                                
                                if merged_ranges_right:
                                    print("右侧合并单元格信息:")
                                    # 记录右侧的合并单元格的信息（删除列前记录）
                                    for rng in merged_ranges_right:
                                        top_left = sheet.cell(rng.min_row, rng.min_col)
                                        value = top_left.value if top_left else None
                                        print(f"  合并区域: {rng.coord}, 值: '{value}'") 

                                        # 记录合并信息
                                        merge_info_list.append({
                                            'min_row': rng.min_row,
                                            'max_row': rng.max_row,
                                            'min_col': rng.min_col,
                                            'max_col': rng.max_col,
                                            'value': value,
                                            'range_str': str(rng)
                                        })                         

                            # 模式为"ignore"时，要处理交集和删除列右侧的单元格
                            elif merged_ranges and merge_mode == 'ignore':    
                                print(f"检测到 {len(merged_ranges)} 个受影响的(包括交集和右侧)合并单元格")
                                
                                # 记录所有合并区域的信息（删除列前记录）
                                for rng in merged_ranges:
                                    top_left = sheet.cell(rng.min_row, rng.min_col)
                                    value = top_left.value if top_left else None
                                    print(f"  合并区域: {rng.coord}, 值: '{value}'")
                                    
                                    # 记录合并信息
                                    merge_info_list.append({
                                        'min_row': rng.min_row,
                                        'max_row': rng.max_row,
                                        'min_col': rng.min_col,
                                        'max_col': rng.max_col,
                                        'value': value,
                                        'range_str': str(rng)
                                    })

                            # 根据已经已载入不同值的merge_info_list重建合并单元格
                            if merge_info_list:    
                                # 1. 先解除所有合并
                                print("第一步: 解除所有合并")
                                for info in merge_info_list:
                                    try:
                                        if info['range_str'] in [str(m) for m in sheet.merged_cells]:
                                            sheet.unmerge_cells(info['range_str'])
                                            print(f"  ✓ 已解除: {info['range_str']}")
                                    except Exception as e:
                                        print(f"  ⚠ 解除警告: {str(e)}")
                                
                                # 2. 执行删除列操作（安全删除——从右往左删除）
                                print("\n第二步: 删除列")
                                columns_to_delete = list(range(position, position + count))
                                for col_idx in sorted(columns_to_delete, reverse=True):
                                    print(f"  删除列: {col_idx}")
                                    sheet.delete_cols(col_idx, 1)
                                
                                # 3. 重建所有合并单元格
                                print("\n第三步: 重建合并区域")
                                for info in merge_info_list:
                                    # 计算新的列范围
                                    new_min_col = info['min_col'] - len([c for c in columns_to_delete if c < info['min_col']])
                                    new_max_col = info['max_col'] - len([c for c in columns_to_delete if c <= info['max_col']]) 
                                    
                                    if any(c == info['min_col'] for c in columns_to_delete):
                                        print(f"  × 合并区域左边界为删除列，无需合并和恢复值")
                                        continue                                    
                                    # 检查新位置是否有效
                                    if new_min_col < 1 or new_max_col < 1 or new_min_col > new_max_col:
                                        print(f"  × 跳过无效范围: 原范围为：MIN：{info['min_col']} ～ MAX：{info['max_col']} => 新范围为：MIN：{new_min_col} ～ MAX：{new_max_col}")
                                        continue

                                    if  new_min_col == new_max_col and info['min_row'] == info['max_row']:
                                        print(f"  ✓ 删除后只有一个单元格，无需合并，原来的值为：{info['value']}")
                                        continue 
                                                                      
                                    # 有效合并范围执行合并操作
                                    if new_min_col >= 1 and new_max_col >= 1 and new_min_col < new_max_col:                                    
                                        # 创建新的合并范围
                                        new_range_str = (
                                            f"{get_column_letter(new_min_col)}{info['min_row']}:"
                                            f"{get_column_letter(new_max_col)}{info['max_row']}"
                                        )
                                        
                                        try:
                                            # 创建新的合并区域
                                            sheet.merge_cells(new_range_str)
                                            print(f"  ✓ 创建新合并: {new_range_str}")

                                            # 设置左上角单元格的值
                                            sheet.cell(info['min_row'], new_min_col).value = info['value']                                            
                                            
                                            print(f"  ✓ 值已恢复: '{info['value']}'")
                                        
                                        except Exception as e:
                                            print(f"  ⚠ 创建合并失败: {str(e)}")
                                
                                print("合并区域重建完成")      

                            # 如果都没有交集或者右侧合并单元格需要处理，则直接删除列
                            else:
                                # 直接执行————删除列操作（安全删除——从右往左删除）
                                print("\n直接删除列：")
                                columns_to_delete = list(range(position, position + count))
                                for col_idx in sorted(columns_to_delete, reverse=True):
                                    print(f"  删除列: {col_idx}")
                                    sheet.delete_cols(col_idx, 1)

                        except ValueError as ve:
                            # 提供更具体的错误信息
                            raise ValueError(f"删除列失败: {str(ve)}")
                        except Exception as e:
                            # 捕获其他可能的错误
                            raise ValueError(f"删除列时发生错误: {str(e)}")
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中删除列失败: {str(e)}")
                raise
    
    def hide_rows(self, file_paths, sheet_indexes, position, count=1):
        """
        隐藏行
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            position: 隐藏位置（行号）
            count: 隐藏行数
        """
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                for sheet_index in sheet_indexes:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        
                        # 隐藏指定范围的行
                        for row in range(position, position + count):
                            sheet.row_dimensions[row].hidden = True
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中隐藏行失败: {str(e)}")
                raise
    
    def unhide_rows(self, file_paths, sheet_indexes, position, count=1):
        """
        显示行
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            position: 显示位置（行号）
            count: 显示行数
        """
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                for sheet_index in sheet_indexes:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        
                        # 显示指定范围的行
                        for row in range(position, position + count):
                            sheet.row_dimensions[row].hidden = False
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中显示行失败: {str(e)}")
                raise
    
    def hide_columns(self, file_paths, sheet_indexes, position, count=1):
        """
        隐藏列
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            position: 隐藏位置（列号）
            count: 隐藏列数
        """
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                for sheet_index in sheet_indexes:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        
                        # 隐藏指定范围的列
                        for col in range(position, position + count):
                            col_letter = get_column_letter(col)
                            sheet.column_dimensions[col_letter].hidden = True
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中隐藏列失败: {str(e)}")
                raise
    
    def unhide_columns(self, file_paths, sheet_indexes, position, count=1):
        """
        显示列
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            position: 显示位置（列号）
            count: 显示列数
        """
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                for sheet_index in sheet_indexes:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        
                        # 显示指定范围的列
                        for col in range(position, position + count):
                            col_letter = get_column_letter(col)
                            sheet.column_dimensions[col_letter].hidden = False
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中显示列失败: {str(e)}")
                raise
    
    def delete_hidden_rows(self, file_paths, sheet_indexes, merge_mode='ignore'):
        """
        删除所有隐藏的行
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            merge_mode: 合并单元格处理模式 ('ignore', 'unmerge_only', 'unmerge_keep_value')
            
        Returns:
            dict: 删除结果详情，包含每个文件每个工作表删除的行信息
        """
        delete_results = {}
        
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                file_results = {}
                
                for sheet_index in sheet_indexes:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        sheet_name = wb.sheetnames[sheet_index]
                        
                        # 收集所有隐藏的行号 - 改进检测逻辑
                        hidden_rows = []
                        
                        # 检查所有可能的行，包括超出max_row的行
                        max_check_row = max(sheet.max_row, 1000)  # 至少检查到1000行
                        
                        for row_num in range(1, max_check_row + 1):
                            # 检查行维度是否存在且隐藏
                            if row_num in sheet.row_dimensions:
                                if sheet.row_dimensions[row_num].hidden:
                                    hidden_rows.append(row_num)
                            else:
                                # 对于不在row_dimensions中的行，检查是否有隐藏属性
                                # 创建行维度对象来检查
                                row_dim = sheet.row_dimensions[row_num]
                                if hasattr(row_dim, 'hidden') and row_dim.hidden:
                                    hidden_rows.append(row_num)
                        
                        # 如果没有隐藏行，跳过
                        if not hidden_rows:
                            file_results[sheet_name] = "无隐藏行"
                            continue
                        
                        # 记录删除前的行号
                        original_hidden_rows = hidden_rows.copy()
                        print("所有的隐藏行:", original_hidden_rows)

                        # 每次删除一行
                        Row_Height = 1

                        # 每一行调用delete_rows进行安全删除——从下往上删除
                        hidden_rows_sorted = sorted(hidden_rows, reverse=True)  # 如果是块格式则用key=lambda x: x[1]
                        for row in hidden_rows_sorted:
                            # 先取消隐藏属性
                            sheet.row_dimensions[row].hidden = False
                            # 再删除行
                            self.delete_rows(file_paths, sheet_indexes, row, Row_Height, merge_mode)
                            print(f"  成功删除行：{row}")
                        
                        # 记录删除结果
                        if original_hidden_rows:
                            file_results[sheet_name] = f"删除了第 {', '.join(map(str, original_hidden_rows))} 行（共{len(original_hidden_rows)}行）"
                        else:
                            file_results[sheet_name] = "无隐藏行"
                
                delete_results[Path(file_path).name] = file_results
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中删除隐藏行失败: {str(e)}")
                delete_results[Path(file_path).name] = f"删除失败: {str(e)}"
                raise
        
        return delete_results
    
    def delete_hidden_columns(self, file_paths, sheet_indexes, merge_mode='ignore'):
        """
        删除所有隐藏的列
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表
            merge_mode: 合并单元格处理模式 ('ignore', 'unmerge_only', 'unmerge_keep_value')
            
        Returns:
            dict: 删除结果详情，包含每个文件每个工作表删除的列信息
        """
        delete_results = {}
        
        for file_path in file_paths:
            try:
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                file_results = {}
                
                for sheet_index in sheet_indexes:
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                        sheet_name = wb.sheetnames[sheet_index]
                        
                        # 收集所有隐藏的列号 - 改进检测逻辑
                        hidden_columns = []
                        
                        # 检查所有可能的列，包括超出max_column的列
                        max_check_col = max(sheet.max_column, 100)  # 至少检查到100列（CV列）
                        
                        for col_num in range(1, max_check_col + 1):
                            col_letter = get_column_letter(col_num)
                            # 检查列维度是否存在且隐藏
                            if col_letter in sheet.column_dimensions:
                                if sheet.column_dimensions[col_letter].hidden:
                                    hidden_columns.append(col_num)
                            else:
                                # 对于不在column_dimensions中的列，检查是否有隐藏属性
                                # 创建列维度对象来检查
                                col_dim = sheet.column_dimensions[col_letter]
                                if hasattr(col_dim, 'hidden') and col_dim.hidden:
                                    hidden_columns.append(col_num)
                        
                        # 如果没有隐藏列，跳过
                        if not hidden_columns:
                            file_results[sheet_name] = "无隐藏列"
                            continue
                        
                        # 记录删除前的列号
                        original_hidden_columns = hidden_columns.copy()
                        # 记录删除前的列字母
                        hidden_col_letters = [get_column_letter(col) for col in original_hidden_columns]
                        print("所有的隐藏列:", hidden_col_letters)

                        # 每次删除一列
                        Col_Width = 1

                        # 每一列调用delete_columns进行安全删除——从右往左删除
                        hidden_columns_sorted = sorted(hidden_columns, reverse=True)
                        for col in hidden_columns_sorted:
                            # 先取消隐藏属性
                            sheet.column_dimensions[get_column_letter(col)].hidden = False
                            # 再删除列
                            self.delete_columns(file_paths, sheet_indexes, col, Col_Width, merge_mode)
                            print(f"  成功删除列：{get_column_letter(col)}")
                        
                        # 记录删除结果
                        if original_hidden_columns:
                            file_results[sheet_name] = f"删除了第 {', '.join(hidden_col_letters)} 列（共{len(original_hidden_columns)}列）"
                        else:
                            file_results[sheet_name] = "无隐藏列"
                
                delete_results[Path(file_path).name] = file_results
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中删除隐藏列失败: {str(e)}")
                delete_results[Path(file_path).name] = f"删除失败: {str(e)}"
                raise
        
        return delete_results
    
    def create_worksheet(self, file_paths, sheet_name):
        """
        创建工作表
        
        Args:
            file_paths: Excel文件路径列表
            sheet_name: 要创建的工作表名称
        """
        for file_path in file_paths:
            try:
                # 确保使用工作簿副本，而不是直接修改原始工作簿
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                # 创建工作表
                wb.create_sheet(title=sheet_name)
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                
            except Exception as e:
                print(f"在文件 {file_path} 中创建工作表失败: {str(e)}")
                raise
    
    def delete_worksheet(self, file_paths, sheet_name):
        """
        删除工作表
        
        Args:
            file_paths: Excel文件路径列表
            sheet_name: 要删除的工作表名称
        """
        for file_path in file_paths:
            try:
                # 确保使用工作簿副本，而不是直接修改原始工作簿
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                # 检查工作表是否存在
                if sheet_name in wb.sheetnames:
                    # 删除工作表
                    del wb[sheet_name]
                    
                    # 保存修改后的工作簿到临时文件
                    temp_path = self.temp_files.get(file_path)
                    if temp_path:
                        wb.save(temp_path)
                else:
                    print(f"工作表 {sheet_name} 在文件 {file_path} 中不存在")
                    continue                
                
            except Exception as e:
                print(f"在文件 {file_path} 中删除工作表失败: {str(e)}")
                raise
    
    def _get_sheet_by_index_or_name(self, wb, sheet_index_or_name, file_path=None):
        """
        根据索引或名称获取工作表
        
        Args:
            wb: 工作簿对象
            sheet_index_or_name: 工作表索引或名称
            file_path: 文件路径，用于检查是否在选定的工作表中
            
        Returns:
            工作表对象
        """
        if isinstance(sheet_index_or_name, int):
            # 按索引获取
            if 0 <= sheet_index_or_name < len(wb.sheetnames):
                sheet_name = wb.sheetnames[sheet_index_or_name]
                
                # 检查是否在选定的工作表中
                if file_path and self.selected_worksheets and file_path in self.selected_worksheets:
                    if sheet_name not in self.selected_worksheets[file_path]:
                        raise ValueError(f"工作表 '{sheet_name}' 不在选定的工作表列表中")
                
                return wb[sheet_name]
            else:
                raise ValueError(f"工作表索引 {sheet_index_or_name} 超出范围")
        else:
            # 按名称获取
            if sheet_index_or_name in wb.sheetnames:
                # 检查是否在选定的工作表中
                if file_path and self.selected_worksheets and file_path in self.selected_worksheets:
                    if sheet_index_or_name not in self.selected_worksheets[file_path]:
                        raise ValueError(f"工作表 '{sheet_index_or_name}' 不在选定的工作表列表中")
                
                return wb[sheet_index_or_name]
            else:
                raise ValueError(f"工作表 '{sheet_index_or_name}' 不存在")
                
    def merge_cells(self, file_paths, sheet_indexes, range_str):
        """
        合并单元格，始终保留左上角单元格的值
        
        与Excel默认行为一致，当指定范围内有多个单元格包含值时，
        合并后只保留左上角单元格的值，其他单元格的值将被丢弃。
        
        Args:
            file_paths: Excel文件路径列表
            sheet_indexes: 工作表索引列表，可以是整数索引或工作表名称
            range_str: 要合并的单元格范围，例如'A1:B3'
        
        Raises:
            ValueError: 当工作表索引无效时
            Exception: 当合并操作失败时
        """
        from openpyxl.utils import range_boundaries
        
        for file_path in file_paths:
            try:
                # 确保使用工作簿副本，而不是直接修改原始工作簿
                wb = self.workbooks.get(file_path)
                if not wb:
                    # 创建临时文件副本
                    temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
                    shutil.copy2(file_path, temp_path)
                    self.temp_files[file_path] = temp_path
                    
                    # 加载临时文件的工作簿
                    wb = openpyxl.load_workbook(temp_path)
                    self.workbooks[file_path] = wb
                
                for sheet_index in sheet_indexes:
                    # 支持通过索引或名称访问工作表
                    if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_index]]
                    elif isinstance(sheet_index, str) and sheet_index in wb.sheetnames:
                        sheet = wb[sheet_index]
                    else:
                        raise ValueError(f"无效的工作表索引或名称: {sheet_index}")
                        
                    # 获取合并范围的边界
                    min_col, min_row, max_col, max_row = range_boundaries(range_str)
                    
                    # 保存左上角单元格的值
                    top_left_value = sheet.cell(row=min_row, column=min_col).value
                    
                    # 检查范围内是否有其他非空值
                    has_other_values = False
                    non_empty_cells = []
                    
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            if row == min_row and col == min_col:
                                continue  # 跳过左上角单元格
                            cell = sheet.cell(row=row, column=col)
                            if cell.value is not None:
                                has_other_values = True
                                cell_addr = f"{get_column_letter(col)}{row}"
                                non_empty_cells.append(cell_addr)
                    
                    # 合并单元格前记录警告信息
                    if has_other_values:
                        cells_str = ", ".join(non_empty_cells[:5])
                        if len(non_empty_cells) > 5:
                            cells_str += f" 等共 {len(non_empty_cells)} 个单元格"
                        print(f"警告: 在文件 {file_path} 的工作表 {sheet.title} 中，"
                              f"合并范围 {range_str} 内存在多个非空值({cells_str})，"
                              f"仅保留左上角单元格 {range_str.split(':')[0]} 的值。")
                    
                    # 合并单元格
                    sheet.merge_cells(range_str)
                    
                    # 确保合并后的单元格保留左上角的值
                    merged_cell = sheet.cell(row=min_row, column=min_col)
                    merged_cell.value = top_left_value
                
                # 保存修改后的工作簿到临时文件
                temp_path = self.temp_files.get(file_path)
                if temp_path:
                    wb.save(temp_path)
                else:
                    # 如果没有临时文件路径，这是一个错误情况
                    raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
                
            except Exception as e:
                print(f"在文件 {file_path} 中合并单元格失败: {str(e)}")
                raise
    
    def close_workbooks(self):
        """
        关闭所有工作簿
        """
        for wb in self.workbooks.values():
            try:
                wb.close()
            except:
                pass
        self.workbooks.clear()