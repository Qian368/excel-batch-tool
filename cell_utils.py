#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
单元格工具模块
提供处理单元格范围和合并单元格的通用方法
"""

import re
from openpyxl.utils import get_column_letter, column_index_from_string


def parse_cell_range(range_str):
    """
    解析单元格范围字符串，支持单个单元格、范围和多个范围
    
    Args:
        range_str: 单元格范围字符串，如 "A1", "A1:B5", "A1,C3:D4"
        
    Returns:
        list: 单元格范围列表，每个元素为 (start_cell, end_cell) 元组
    """
    # 替换中文符号
    range_str = range_str.replace('，', ',').replace('：', ':')
    
    # 分割多个范围
    ranges = range_str.split(',')
    result = []
    
    for r in ranges:
        r = r.strip()
        if ':' in r:
            # 处理范围 (如 A1:B5)
            start, end = r.split(':')
            result.append((start.strip(), end.strip()))
        else:
            # 处理单个单元格 (如 A1)
            result.append((r, r))
    
    return result


def get_cell_coordinates(cell_ref):
    """
    将单元格引用转换为行列坐标
    
    Args:
        cell_ref: 单元格引用，如 "A1"
        
    Returns:
        tuple: (行号, 列号) 元组
    """
    match = re.match(r'([A-Za-z]+)(\d+)', cell_ref)
    if not match:
        raise ValueError(f"无效的单元格引用: {cell_ref}")
        
    col, row = match.groups()
    return int(row), column_index_from_string(col)


def process_cell_ranges(sheet, cell_ranges, process_func):
    """
    处理指定的单元格范围，并应用处理函数
    
    Args:
        sheet: 工作表对象
        cell_ranges: 单元格范围列表，每个元素为 (start_cell, end_cell) 元组
        process_func: 处理函数，接收单元格对象作为参数
        
    Returns:
        dict: 包含操作结果的字典
    """
    try:
        # 记录已处理的单元格坐标，避免重复处理
        processed_cells = set()
        
        for start_cell, end_cell in cell_ranges:
            # 处理单个单元格
            if start_cell == end_cell:
                # 获取单元格坐标
                row, col = get_cell_coordinates(start_cell)
                cell_coord = (row, col)
                
                # 如果单元格已处理过，跳过
                if cell_coord in processed_cells:
                    continue
                
                # 直接处理该单元格
                cell = sheet.cell(row=row, column=col)
                process_func(cell)
                # 标记该单元格已处理
                processed_cells.add(cell_coord)
            else:
                # 处理单元格范围
                # 获取起始和结束单元格的坐标
                start_row, start_col = get_cell_coordinates(start_cell)
                end_row, end_col = get_cell_coordinates(end_cell)
                
                # 确保起始坐标小于等于结束坐标
                min_row = min(start_row, end_row)
                max_row = max(start_row, end_row)
                min_col = min(start_col, end_col)
                max_col = max(start_col, end_col)
                
                # 直接处理用户指定的范围，不扩展到合并单元格
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell_coord = (row, col)
                        # 如果单元格已处理过，跳过
                        if cell_coord in processed_cells:
                            continue
                        
                        # 处理单元格
                        cell = sheet.cell(row=row, column=col)
                        process_func(cell)
                        # 标记该单元格已处理
                        processed_cells.add(cell_coord)
                        
        return {'success': True}
    except Exception as e:
        return {'success': False, 'message': str(e)}


def get_merged_cell_value(sheet, cell_position):
    """
    获取单元格的值，如果是合并单元格，则返回左上角单元格的值
    
    Args:
        sheet: 工作表对象
        cell_position: 单元格位置，如 "A1"
        
    Returns:
        单元格的值
    """
    # 获取单元格坐标
    row, col = get_cell_coordinates(cell_position)
    
    # 检查该单元格是否在合并单元格内
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and 
            merged_range.min_col <= col <= merged_range.max_col):
            # 如果单元格在合并单元格内，返回左上角单元格的值
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    
    # 如果不在合并单元格内，返回该单元格的值
    return sheet.cell(row=row, column=col).value