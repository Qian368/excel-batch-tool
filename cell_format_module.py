#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
单元格格式化模块
提供Excel单元格格式化相关功能
"""

from openpyxl.styles import Font, PatternFill, Border, Side
from cell_utils import parse_cell_range, process_cell_ranges
from excel_utils import ensure_workbook, get_sheets_to_process


# 颜色映射表
COLOR_MAP = {
    '红色': 'FF0000',
    '绿色': '00FF00',
    '蓝色': '0000FF',
    '黑色': '000000',
    '白色': 'FFFFFF',
    '黄色': 'FFFF00',
    '紫色': '800080',
    '橙色': 'FFA500',
    '灰色': '808080'
}


def change_font_color(processor, file_paths, color, range_mode='specific', range_str=None):
    """
    修改单元格字体颜色
    
    Args:
        processor: ExcelProcessor实例
        file_paths: 要处理的Excel文件路径列表
        color: 字体颜色名称
        range_mode: 范围模式，'specific'表示指定范围，'entire_sheet'表示整个工作表
        range_str: 当range_mode为'specific'时，指定要处理的单元格范围
        
    Returns:
        dict: 包含操作结果的字典
    """
    # 获取RGB颜色代码
    rgb = COLOR_MAP.get(color, '000000')  # 默认黑色
    
    # 定义应用字体颜色的函数
    def apply_font_color(cell, rgb_color=rgb):
        if cell.font:
            # 保留原有字体属性，只修改颜色
            new_font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=rgb_color
            )
            cell.font = new_font
        else:
            cell.font = Font(color=rgb_color)
    
    try:
        for file_path in file_paths:
            # 使用通用函数确保工作簿已加载
            wb, temp_path = ensure_workbook(processor, file_path)
            
            # 使用通用函数获取要处理的工作表
            sheets_to_process = get_sheets_to_process(processor, file_path)
            
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]
                
                if range_mode == 'entire_sheet':
                    # 修改整个工作表的字体颜色
                    for row in sheet.iter_rows():
                        for cell in row:
                            apply_font_color(cell)
                else:  # specific
                    if not range_str:
                        raise ValueError("未指定单元格范围")
                    
                    # 处理中文符号
                    range_str = range_str.replace('，', ',').replace('：', ':')
                    
                    # 解析单元格范围
                    cell_ranges = parse_cell_range(range_str)
                    
                    # 直接使用用户输入的范围，不求并集
                    result = process_cell_ranges(sheet, cell_ranges, apply_font_color)
                    if not result['success']:
                        raise ValueError(f"修改字体颜色失败: {result['message']}")
            
            # 保存修改后的工作簿到临时文件
            if temp_path:
                wb.save(temp_path)
            else:
                # 如果没有临时文件路径，这是一个错误情况
                raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
        
        return {
            'success': True,
            'message': f"修改字体颜色（颜色：{color}）执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"执行失败：{str(e)}"
        }


def change_fill_color(processor, file_paths, color, range_mode='specific', range_str=None):
    """
    修改单元格填充颜色
    
    Args:
        processor: ExcelProcessor实例
        file_paths: 要处理的Excel文件路径列表
        color: 填充颜色名称
        range_mode: 范围模式，'specific'表示指定范围，'entire_sheet'表示整个工作表
        range_str: 当range_mode为'specific'时，指定要处理的单元格范围
        
    Returns:
        dict: 包含操作结果的字典
    """
    # 获取RGB颜色代码
    rgb = COLOR_MAP.get(color, '000000')  # 默认黑色
    
    # 定义应用填充颜色的函数
    def apply_fill_color(cell, rgb_color=rgb):
        cell.fill = PatternFill(start_color=rgb_color, end_color=rgb_color, fill_type='solid')
    
    try:
        for file_path in file_paths:
            # 使用通用函数确保工作簿已加载
            wb, temp_path = ensure_workbook(processor, file_path)
            
            # 使用通用函数获取要处理的工作表
            sheets_to_process = get_sheets_to_process(processor, file_path)
            
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]
                
                if range_mode == 'entire_sheet':
                    # 修改整个工作表的填充颜色
                    for row in sheet.iter_rows():
                        for cell in row:
                            apply_fill_color(cell)
                else:  # specific
                    if not range_str:
                        raise ValueError("未指定单元格范围")
                    
                    # 处理中文符号
                    range_str = range_str.replace('，', ',').replace('：', ':')
                    
                    # 解析单元格范围
                    cell_ranges = parse_cell_range(range_str)
                    
                    # 直接使用用户输入的范围，不求并集
                    result = process_cell_ranges(sheet, cell_ranges, apply_fill_color)
                    if not result['success']:
                        raise ValueError(f"修改填充颜色失败: {result['message']}")
            
            # 保存修改后的工作簿到临时文件
            if temp_path:
                wb.save(temp_path)
            else:
                # 如果没有临时文件路径，这是一个错误情况
                raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
        
        return {
            'success': True,
            'message': f"修改填充颜色（颜色：{color}）执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"执行失败：{str(e)}"
        }


def add_border(processor, file_paths, range_str, border_style='thin'):
    """
    添加单元格边框
    
    Args:
        processor: ExcelProcessor实例
        file_paths: 要处理的Excel文件路径列表
        range_str: 指定要处理的单元格范围
        border_style: 边框样式，默认为'thin'
        
    Returns:
        dict: 包含操作结果的字典
    """
    try:
        # 创建边框样式
        border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style)
        )
        
        # 定义应用边框的函数
        def apply_border(cell):
            cell.border = border
        
        for file_path in file_paths:
            # 使用通用函数确保工作簿已加载
            wb, temp_path = ensure_workbook(processor, file_path)
            
            # 使用通用函数获取要处理的工作表
            sheets_to_process = get_sheets_to_process(processor, file_path)
            
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]
                
                if not range_str:
                    raise ValueError("未指定单元格范围")
                
                # 处理中文符号
                range_str = range_str.replace('，', ',').replace('：', ':')
                
                # 解析单元格范围
                cell_ranges = parse_cell_range(range_str)
                
                # 直接使用用户输入的范围，不求并集
                result = process_cell_ranges(sheet, cell_ranges, apply_border)
                if not result['success']:
                    raise ValueError(f"添加边框失败: {result['message']}")
            
            # 保存修改后的工作簿到临时文件
            if temp_path:
                wb.save(temp_path)
            else:
                # 如果没有临时文件路径，这是一个错误情况
                raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
        
        return {
            'success': True,
            'message': f"添加边框执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"执行失败：{str(e)}"
        }


def remove_border(processor, file_paths, range_str):
    """
    移除单元格边框
    
    Args:
        processor: ExcelProcessor实例
        file_paths: 要处理的Excel文件路径列表
        range_str: 指定要处理的单元格范围
        
    Returns:
        dict: 包含操作结果的字典
    """
    try:
        # 创建无边框样式
        no_border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None)
        )
        
        # 定义移除边框的函数
        def remove_cell_border(cell):
            cell.border = no_border
        
        for file_path in file_paths:
            # 使用通用函数确保工作簿已加载
            wb, temp_path = ensure_workbook(processor, file_path)
            
            # 使用通用函数获取要处理的工作表
            sheets_to_process = get_sheets_to_process(processor, file_path)
            
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]
                
                if not range_str:
                    raise ValueError("未指定单元格范围")
                
                # 处理中文符号
                range_str = range_str.replace('，', ',').replace('：', ':')
                
                # 解析单元格范围
                cell_ranges = parse_cell_range(range_str)
                
                # 使用通用方法处理单元格范围，包括合并单元格
                result = process_cell_ranges(sheet, cell_ranges, remove_cell_border)
                if not result['success']:
                    raise ValueError(f"移除边框失败: {result['message']}")
            
            # 保存修改后的工作簿到临时文件
            if temp_path:
                wb.save(temp_path)
            else:
                # 如果没有临时文件路径，这是一个错误情况
                raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
        
        return {
            'success': True,
            'message': f"移除边框执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"执行失败：{str(e)}"
        }


def modify_cell_content(processor, file_paths, position, content):
    """
    修改单元格内容
    
    Args:
        processor: ExcelProcessor实例
        file_paths: 要处理的Excel文件路径列表
        position: 单元格位置，如 "A1" 或 "A1:B2,C3"
        content: 要设置的内容
        
    Returns:
        dict: 包含操作结果的字典
    """
    try:
        if not position:
            raise ValueError("未指定单元格位置")
        
        # 处理中文符号
        position = position.replace('，', ',').replace('：', ':')
        
        # 解析单元格范围
        cell_ranges = parse_cell_range(position)
        
        # 定义修改单元格内容的函数
        def set_cell_content(cell):
            try:
                # 检查该单元格是否在合并单元格内
                row, col = cell.row, cell.column
                target_cell = cell  # 默认目标单元格就是当前单元格
                
                # 遍历所有合并单元格范围
                for merged_range in cell.parent.merged_cells.ranges:
                    if (merged_range.min_row <= row <= merged_range.max_row and 
                        merged_range.min_col <= col <= merged_range.max_col):
                        # 如果单元格在合并单元格内，设置左上角单元格的值
                        target_cell = cell.parent.cell(row=merged_range.min_row, column=merged_range.min_col)
                        break
                
                # 设置目标单元格的值
                target_cell.value = content
            except Exception as e:
                # 记录异常但不中断整个操作
                print(f"设置单元格 {cell.coordinate} 的值时出错: {str(e)}")
        
        for file_path in file_paths:
            # 使用通用函数确保工作簿已加载
            wb, temp_path = ensure_workbook(processor, file_path)
            
            # 使用通用函数获取要处理的工作表
            sheets_to_process = get_sheets_to_process(processor, file_path)
            
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]
                
                # 直接使用用户输入的范围，不求并集
                result = process_cell_ranges(sheet, cell_ranges, set_cell_content)
                if not result['success']:
                    raise ValueError(f"修改单元格内容失败: {result['message']}")
            
            # 保存修改后的工作簿到临时文件
            if temp_path:
                wb.save(temp_path)
            else:
                # 如果没有临时文件路径，这是一个错误情况
                raise ValueError(f"找不到文件 {file_path} 的临时文件路径")
        
        return {
            'success': True,
            'message': f"修改单元格内容（单元格：{position}）执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"执行失败：{str(e)}"
        }