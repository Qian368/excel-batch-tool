#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
单元格格式操作模块
提供单元格字体颜色、填充颜色、边框和内容修改等功能
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import re

from messages import (
    OPERATION_FONT_COLOR, OPERATION_FILL_COLOR, 
    OPERATION_ADD_BORDER, OPERATION_REMOVE_BORDER, OPERATION_MODIFY_CONTENT,
    RANGE_DESC_SPECIFIC, RANGE_DESC_ENTIRE_SHEET
)

from cell_utils import parse_cell_range, process_cell_ranges, get_merged_cell_value


# 使用cell_utils.py中的parse_cell_range函数，此处不再需要重复定义


def get_color_rgb(color_name):
    """
    根据颜色名称获取RGB值
    
    Args:
        color_name: 颜色名称
        
    Returns:
        str: RGB颜色值，格式为RRGGBB
    """
    color_map = {
        '红色': 'FF0000',
        '绿色': '00FF00',
        '蓝色': '0000FF',
        '黑色': '000000',
        '白色': 'FFFFFF',
        '黄色': 'FFFF00',
        '紫色': '800080',
        '橙色': 'FFA500',
        '灰色': '808080'
    }
    return color_map.get(color_name, '000000')  # 默认黑色


def change_font_color(workbook, params):
    """
    修改单元格字体颜色
    
    Args:
        workbook: openpyxl工作簿对象
        params: 包含操作参数的字典，需要包含color和range_mode，
                如果range_mode为'specific'，还需要包含range_str
                
    Returns:
        dict: 包含操作结果的字典
    """
    try:
        color_name = params.get('color', '黑色')
        range_mode = params.get('range_mode', 'specific')
        rgb = get_color_rgb(color_name)
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if range_mode == 'entire_sheet':
                # 修改整个工作表的字体颜色
                for row in sheet.iter_rows():
                    for cell in row:
                        apply_font_color(cell, rgb)
            else:  # specific
                # 修改指定范围的字体颜色
                range_str = params.get('range_str', '')
                if not range_str:
                    return {
                        'success': False,
                        'message': f"{OPERATION_FONT_COLOR} 执行失败: 未指定单元格范围"
                    }
                
                # 解析单元格范围
                try:
                    cell_ranges = parse_cell_range(range_str)
                except Exception as e:
                    return {
                        'success': False,
                        'message': f"{OPERATION_FONT_COLOR} 执行失败: 单元格范围格式错误 - {str(e)}"
                    }
                
                # 直接使用用户输入的范围，不求并集
                from cell_utils import process_cell_ranges
                result = process_cell_ranges(sheet, cell_ranges, lambda cell: apply_font_color(cell, rgb))
                if not result['success']:
                    return {
                        'success': False,
                        'message': f"{OPERATION_FONT_COLOR} 执行失败: {result['message']}"
                    }
        
        # 构建成功消息
        range_desc = RANGE_DESC_ENTIRE_SHEET if range_mode == 'entire_sheet' else f"{RANGE_DESC_SPECIFIC}: {params.get('range_str', '')}"
        return {
            'success': True,
            'message': f"{OPERATION_FONT_COLOR}（颜色：{color_name}，{range_desc}）执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"{OPERATION_FONT_COLOR} 执行失败: {str(e)}"
        }


def apply_font_color(cell, rgb):
    """
    应用字体颜色到单元格，安全处理合并单元格
    
    Args:
        cell: 单元格对象
        rgb: 颜色RGB值
    """
    try:
        # 检查是否是MergedCell对象
        from openpyxl.cell.cell import MergedCell
        if isinstance(cell, MergedCell):
            return
        
        # 创建字体对象
        if cell.font:
            # 保留原有字体属性，只修改颜色
            new_font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=rgb
            )
            cell.font = new_font
        else:
            cell.font = Font(color=rgb)
    except AttributeError as e:
        if "read-only" in str(e):
            return
        else:
            raise e
    except Exception as e:
        print(f"设置单元格 {cell.coordinate} 的字体颜色时出错: {str(e)}")


def apply_fill_color(cell, fill):
    """
    应用填充颜色到单元格，安全处理合并单元格
    
    Args:
        cell: 单元格对象
        fill: 填充对象
    """
    try:
        # 检查是否是MergedCell对象
        from openpyxl.cell.cell import MergedCell
        if isinstance(cell, MergedCell):
            return
        
        cell.fill = fill
    except AttributeError as e:
        if "read-only" in str(e):
            return
        else:
            raise e
    except Exception as e:
        print(f"设置单元格 {cell.coordinate} 的填充颜色时出错: {str(e)}")


def apply_border(cell, border):
    """
    应用边框到单元格，安全处理合并单元格
    
    Args:
        cell: 单元格对象
        border: 边框对象
    """
    try:
        # 检查是否是MergedCell对象
        from openpyxl.cell.cell import MergedCell
        if isinstance(cell, MergedCell):
            return
        
        cell.border = border
    except AttributeError as e:
        if "read-only" in str(e):
            return
        else:
            raise e
    except Exception as e:
        print(f"设置单元格 {cell.coordinate} 的边框时出错: {str(e)}")


def change_fill_color(workbook, params):
    """
    修改单元格填充颜色
    
    Args:
        workbook: openpyxl工作簿对象
        params: 包含操作参数的字典，需要包含color和range_mode，
                如果range_mode为'specific'，还需要包含range_str
                
    Returns:
        dict: 包含操作结果的字典
    """
    try:
        color_name = params.get('color', '白色')
        range_mode = params.get('range_mode', 'specific')
        rgb = get_color_rgb(color_name)
        
        # 创建填充对象
        fill = PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if range_mode == 'entire_sheet':
                # 修改整个工作表的填充颜色
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.fill = fill
            else:  # specific
                # 修改指定范围的填充颜色
                range_str = params.get('range_str', '')
                if not range_str:
                    return {
                        'success': False,
                        'message': f"{OPERATION_FILL_COLOR} 执行失败: 未指定单元格范围"
                    }
                
                # 解析单元格范围
                try:
                    cell_ranges = parse_cell_range(range_str)
                except Exception as e:
                    return {
                        'success': False,
                        'message': f"{OPERATION_FILL_COLOR} 执行失败: 单元格范围格式错误 - {str(e)}"
                    }
                
                # 直接使用用户输入的范围，不求并集
                from cell_utils import process_cell_ranges
                result = process_cell_ranges(sheet, cell_ranges, lambda cell: apply_fill_color(cell, fill))
                if not result['success']:
                    return {
                        'success': False,
                        'message': f"{OPERATION_FILL_COLOR} 执行失败: {result['message']}"
                    }
        
        # 构建成功消息
        range_desc = RANGE_DESC_ENTIRE_SHEET if range_mode == 'entire_sheet' else f"{RANGE_DESC_SPECIFIC}: {params.get('range_str', '')}"
        return {
            'success': True,
            'message': f"{OPERATION_FILL_COLOR}（颜色：{color_name}，{range_desc}）执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"{OPERATION_FILL_COLOR} 执行失败: {str(e)}"
        }


def modify_cell_border(workbook, params):
    """
    修改单元格边框（添加或移除）
    
    Args:
        workbook: openpyxl工作簿对象
        params: 包含操作参数的字典，需要包含border_mode和range_mode，
                如果range_mode为'specific'，还需要包含range_str
                
    Returns:
        dict: 包含操作结果的字典
    """
    try:
        border_mode = params.get('border_mode', 'add')
        range_mode = params.get('range_mode', 'specific')
        
        # 创建边框对象
        if border_mode == 'add':
            # 创建实线边框
            side = Side(style='thin')
            border = Border(left=side, right=side, top=side, bottom=side)
            operation_desc = OPERATION_ADD_BORDER
        else:  # remove
            # 创建无边框
            border = Border(left=None, right=None, top=None, bottom=None)
            operation_desc = OPERATION_REMOVE_BORDER
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if range_mode == 'entire_sheet':
                # 修改整个工作表的边框
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.border = border
            else:  # specific
                # 修改指定范围的边框
                range_str = params.get('range_str', '')
                if not range_str:
                    return {
                        'success': False,
                        'message': f"{operation_desc} 执行失败: 未指定单元格范围"
                    }
                
                # 解析单元格范围
                try:
                    cell_ranges = parse_cell_range(range_str)
                except Exception as e:
                    return {
                        'success': False,
                        'message': f"{operation_desc} 执行失败: 单元格范围格式错误 - {str(e)}"
                    }
                
                # 直接使用用户输入的范围，不求并集
                from cell_utils import process_cell_ranges
                result = process_cell_ranges(sheet, cell_ranges, lambda cell: apply_border(cell, border))
                if not result['success']:
                    return {
                        'success': False,
                        'message': f"{operation_desc} 执行失败: {result['message']}"
                    }
        
        # 构建成功消息
        range_desc = RANGE_DESC_ENTIRE_SHEET if range_mode == 'entire_sheet' else f"{RANGE_DESC_SPECIFIC}: {params.get('range_str', '')}"
        return {
            'success': True,
            'message': f"{operation_desc}（{range_desc}）执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"{operation_desc} 执行失败: {str(e)}"
        }


def modify_cell_content(workbook, params):
    """
    修改单元格内容
    
    Args:
        workbook: openpyxl工作簿对象
        params: 包含操作参数的字典，需要包含position和content
                
    Returns:
        dict: 包含操作结果的字典
    """
    try:
        cell_position = params.get('position', '')
        new_content = params.get('content', '')
        
        if not cell_position:
            return {
                'success': False,
                'message': f"{OPERATION_MODIFY_CONTENT} 执行失败: 未指定单元格位置"
            }
        
        # 替换中文符号
        cell_position = cell_position.replace('，', ',').replace('：', ':')
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            try:
                # 解析单元格范围
                cell_ranges = parse_cell_range(cell_position)
                
                # 定义处理单元格内容的函数，正确处理合并单元格
                def set_cell_content(cell):
                    from openpyxl.cell.cell import MergedCell
                    
                    try:
                        # 检查当前单元格是否在合并单元格范围内
                        for merged_range in sheet.merged_cells.ranges:
                            if (merged_range.min_row <= cell.row <= merged_range.max_row and 
                                merged_range.min_col <= cell.column <= merged_range.max_col):
                                # 如果在合并单元格范围内，只对左上角单元格设置值
                                if cell.row == merged_range.min_row and cell.column == merged_range.min_col:
                                    cell.value = new_content
                                # 对于合并单元格范围内的其他单元格，跳过处理
                                return
                        
                        # 如果不在合并单元格范围内，直接设置值
                        cell.value = new_content
                    except Exception as e:
                        # 如果设置失败，可能是因为其他原因，记录但不中断整个操作
                        pass
                
                # 直接使用用户输入的范围，但仍需处理合并单元格
                from cell_utils import process_cell_ranges
                result = process_cell_ranges(sheet, cell_ranges, set_cell_content)
                if not result['success']:
                    return {
                        'success': False,
                        'message': f"{OPERATION_MODIFY_CONTENT} 执行失败: {result['message']}"
                    }
            except Exception as e:
                return {
                    'success': False,
                    'message': f"{OPERATION_MODIFY_CONTENT} 执行失败: 无法修改单元格 {cell_position} - {str(e)}"
                }
        
        return {
            'success': True,
            'message': f"{OPERATION_MODIFY_CONTENT}（单元格：{cell_position}，内容：{new_content}）执行成功"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"{OPERATION_MODIFY_CONTENT} 执行失败: {str(e)}"
        }