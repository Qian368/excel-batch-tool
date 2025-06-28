# -*- coding: utf-8 -*-
"""
合并单元格处理工具模块
提供更精确的合并单元格处理逻辑

注意：此模块中的求并集相关函数已不再使用
现在四个步骤（字体颜色、填充颜色、单元格边框、单元格内容）
直接使用用户输入的单元格范围，不再求并集
"""

from cell_utils import get_cell_coordinates, parse_cell_range
from openpyxl.utils import get_column_letter


def get_intersecting_merged_cells(sheet, cell_ranges):
    """
    获取与指定单元格范围有交集的合并单元格
    
    Args:
        sheet: 工作表对象
        cell_ranges: 单元格范围列表，每个元素为 (start_cell, end_cell) 元组
        
    Returns:
        list: 与指定范围有交集的合并单元格范围列表
    """
    intersecting_merged_cells = []
    
    for cell_range in cell_ranges:
        start_cell, end_cell = cell_range
        
        # 获取范围的边界
        start_row, start_col = get_cell_coordinates(start_cell)
        end_row, end_col = get_cell_coordinates(end_cell)
        
        min_row = min(start_row, end_row)
        max_row = max(start_row, end_row)
        min_col = min(start_col, end_col)
        max_col = max(start_col, end_col)
        
        # 检查每个合并单元格是否与当前范围有交集
        for merged_range in sheet.merged_cells.ranges:
            merged_min_row, merged_min_col = merged_range.min_row, merged_range.min_col
            merged_max_row, merged_max_col = merged_range.max_row, merged_range.max_col
            
            # 检查是否有交集
            if not (merged_max_row < min_row or merged_min_row > max_row or 
                    merged_max_col < min_col or merged_min_col > max_col):
                # 有交集，添加到列表中（避免重复）
                merged_range_str = str(merged_range)
                if merged_range_str not in [str(mr) for mr in intersecting_merged_cells]:
                    intersecting_merged_cells.append(merged_range)
    
    return intersecting_merged_cells


def get_union_ranges(sheet, cell_ranges):
    """
    获取指定单元格范围与相关合并单元格的并集
    
    Args:
        sheet: 工作表对象
        cell_ranges: 单元格范围列表，每个元素为 (start_cell, end_cell) 元组
        
    Returns:
        list: 并集后的单元格范围字符串列表
    """
    # 获取与指定范围有交集的合并单元格
    intersecting_merged_cells = get_intersecting_merged_cells(sheet, cell_ranges)
    
    # 创建所有需要处理的范围列表
    all_ranges = []
    
    # 添加原始范围（转换为字符串格式）
    for start_cell, end_cell in cell_ranges:
        if start_cell == end_cell:
            all_ranges.append(start_cell)
        else:
            all_ranges.append(f"{start_cell}:{end_cell}")
    
    # 添加相关的合并单元格范围
    for merged_range in intersecting_merged_cells:
        start_cell = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
        end_cell = f"{get_column_letter(merged_range.max_col)}{merged_range.max_row}"
        
        if start_cell == end_cell:
            merged_range_str = start_cell
        else:
            merged_range_str = f"{start_cell}:{end_cell}"
        
        all_ranges.append(merged_range_str)
    
    return all_ranges


def process_cell_ranges_with_merged_cells(sheet, cell_ranges, process_func):
    """
    处理单元格范围，正确处理合并单元格
    
    Args:
        sheet: 工作表对象
        cell_ranges: 单元格范围列表
        process_func: 处理函数，接受单元格对象作为参数
        
    Returns:
        dict: 包含成功状态和消息的字典
    """
    try:
        # 获取并集范围
        union_ranges = get_union_ranges(sheet, cell_ranges)
        
        # 跟踪已处理的单元格和合并单元格
        processed_cells = set()
        processed_merged_cells = set()
        
        for cell_range in union_ranges:
            # 检查是否为单个单元格还是范围
            if ':' in cell_range:
                start_cell, end_cell = cell_range.split(':')
            else:
                # 单个单元格的情况
                start_cell = end_cell = cell_range
            
            if start_cell == end_cell:
                # 单个单元格
                row, col = get_cell_coordinates(start_cell)
                cell_coord = (row, col)
                
                # 如果单元格已处理过，跳过
                if cell_coord in processed_cells:
                    continue
                
                # 检查该单元格是否在合并单元格内
                in_merged_cell = False
                for merged_range in sheet.merged_cells.ranges:
                    merged_range_str = str(merged_range)
                    if (merged_range.min_row <= row <= merged_range.max_row and 
                        merged_range.min_col <= col <= merged_range.max_col):
                        # 如果单元格在合并单元格内，且该合并单元格尚未处理过
                        if merged_range_str not in processed_merged_cells:
                            # 只处理合并单元格的左上角单元格
                            top_left_cell = sheet.cell(
                                row=merged_range.min_row, 
                                column=merged_range.min_col
                            )
                            process_func(top_left_cell)
                            
                            # 标记整个合并单元格区域的所有单元格为已处理
                            for r in range(merged_range.min_row, merged_range.max_row + 1):
                                for c in range(merged_range.min_col, merged_range.max_col + 1):
                                    processed_cells.add((r, c))
                            
                            # 标记该合并单元格已处理
                            processed_merged_cells.add(merged_range_str)
                        in_merged_cell = True
                        break
                
                # 如果不在合并单元格内，直接处理该单元格
                if not in_merged_cell:
                    cell = sheet.cell(row=row, column=col)
                    process_func(cell)
                    processed_cells.add(cell_coord)
            
            else:
                # 单元格范围
                start_row, start_col = get_cell_coordinates(start_cell)
                end_row, end_col = get_cell_coordinates(end_cell)
                
                min_row = min(start_row, end_row)
                max_row = max(start_row, end_row)
                min_col = min(start_col, end_col)
                max_col = max(start_col, end_col)
                
                # 处理范围内的所有单元格
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell_coord = (row, col)
                        
                        # 如果单元格已处理过，跳过
                        if cell_coord in processed_cells:
                            continue
                        
                        # 检查该单元格是否在合并单元格内
                        in_merged_cell = False
                        for merged_range in sheet.merged_cells.ranges:
                            merged_range_str = str(merged_range)
                            if (merged_range.min_row <= row <= merged_range.max_row and 
                                merged_range.min_col <= col <= merged_range.max_col):
                                # 如果单元格在合并单元格内，且该合并单元格尚未处理过
                                if merged_range_str not in processed_merged_cells:
                                    # 只处理合并单元格的左上角单元格
                                    top_left_cell = sheet.cell(
                                        row=merged_range.min_row, 
                                        column=merged_range.min_col
                                    )
                                    process_func(top_left_cell)
                                    
                                    # 标记整个合并单元格区域的所有单元格为已处理
                                    for r in range(merged_range.min_row, merged_range.max_row + 1):
                                        for c in range(merged_range.min_col, merged_range.max_col + 1):
                                            processed_cells.add((r, c))
                                    
                                    # 标记该合并单元格已处理
                                    processed_merged_cells.add(merged_range_str)
                                in_merged_cell = True
                                break
                        
                        # 如果不在合并单元格内，直接处理该单元格
                        if not in_merged_cell:
                            cell = sheet.cell(row=row, column=col)
                            process_func(cell)
                            processed_cells.add(cell_coord)
        
        return {'success': True}
    
    except Exception as e:
        return {'success': False, 'message': str(e)}


def is_cell_in_merged_range(sheet, row, col):
    """
    检查单元格是否在合并单元格范围内
    
    Args:
        sheet: 工作表对象
        row: 行号（从1开始）
        col: 列号（从1开始）
        
    Returns:
        tuple: (是否在合并范围内, 合并范围对象或None)
    """
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and 
            merged_range.min_col <= col <= merged_range.max_col):
            return True, merged_range
    return False, None


def get_merged_cell_top_left(sheet, row, col):
    """
    获取合并单元格的左上角单元格
    
    Args:
        sheet: 工作表对象
        row: 行号（从1开始）
        col: 列号（从1开始）
        
    Returns:
        tuple: (左上角行号, 左上角列号) 或 (原行号, 原列号)如果不在合并单元格内
    """
    is_merged, merged_range = is_cell_in_merged_range(sheet, row, col)
    if is_merged:
        return merged_range.min_row, merged_range.min_col
    return row, col