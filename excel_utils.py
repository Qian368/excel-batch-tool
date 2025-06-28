#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel工具模块
提供Excel文件处理的通用工具函数
"""

import os
import shutil
from pathlib import Path

import openpyxl


def ensure_workbook(processor, file_path):
    """
    确保工作簿已加载，如果未加载则创建临时副本并加载
    
    Args:
        processor: ExcelProcessor实例
        file_path: Excel文件路径
        
    Returns:
        tuple: (workbook, temp_path) 工作簿对象和临时文件路径
    """
    # 检查工作簿是否已加载
    wb = processor.workbooks.get(file_path)
    temp_path = processor.temp_files.get(file_path)
    
    if not wb:
        # 创建临时文件副本
        temp_path = Path(file_path).parent / f"temp_{Path(file_path).name}"
        shutil.copy2(file_path, temp_path)
        processor.temp_files[file_path] = temp_path
        
        # 加载临时文件的工作簿
        wb = openpyxl.load_workbook(temp_path)
        processor.workbooks[file_path] = wb
    
    return wb, temp_path


def get_sheets_to_process(processor, file_path):
    """
    确定要处理的工作表列表
    
    Args:
        processor: ExcelProcessor实例
        file_path: Excel文件路径
        
    Returns:
        list: 要处理的工作表名称列表
    """
    wb = processor.workbooks.get(file_path)
    if not wb:
        return []
        
    # 确定要处理的工作表
    if file_path in processor.selected_worksheets and processor.selected_worksheets[file_path]:
        # 如果用户选择了特定的工作表，则只处理这些工作表
        return [sheet_name for sheet_name in processor.selected_worksheets[file_path] 
                if sheet_name in wb.sheetnames]
    else:
        # 否则处理所有工作表
        return wb.sheetnames