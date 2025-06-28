#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
报告生成模块
提供Excel执行结果报告的生成功能
"""

import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from message_utils import format_result_message


def generate_report(step_results, output_dir, file_name="执行报告.xlsx"):
    """
    生成Excel格式的执行报告
    
    Args:
        step_results: 步骤执行结果列表
        output_dir: 输出目录
        file_name: 报告文件名
        
    Returns:
        str: 报告文件路径或报告文件路径列表（当处理多个文件时）
    """
    # 检查是否有文件信息，如果有，则为每个文件生成单独的报告
    file_paths = set()
    for result in step_results:
        if 'file_path' in result:
            file_paths.add(result['file_path'])
    
    # 如果没有文件信息或只有一个文件，生成单个报告
    if not file_paths or len(file_paths) == 1:
        return _generate_single_report(step_results, output_dir, file_name)
    else:
        # 为每个文件生成单独的报告
        report_paths = []
        for file_path in file_paths:
            # 获取文件名作为报告名称的一部分
            file_name_base = os.path.basename(file_path)
            report_file_name = f"执行报告_{file_name_base}.xlsx"
            
            # 筛选该文件的步骤结果
            file_results = [result for result in step_results if result.get('file_path') == file_path]
            
            # 生成该文件的报告
            report_path = _generate_single_report(file_results, output_dir, report_file_name)
            report_paths.append(report_path)
        
        return report_paths


def _generate_single_report(step_results, output_dir, file_name="执行报告.xlsx"):
    """
    生成单个Excel格式的执行报告
    
    Args:
        step_results: 步骤执行结果列表
        output_dir: 输出目录
        file_name: 报告文件名
        
    Returns:
        str: 报告文件路径
    """
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "执行结果报告"
    
    # 操作类型英文到中文的映射
    operation_name_map = {
        "convert_formulas_to_values": "公式转值",
        "process_merged_cells_all": "拆分所有合并单元格",
        "process_merged_cells_specific": "拆分指定范围合并单元格",
        "merge_cells": "合并单元格",
        "create_worksheet": "新建工作表",
        "delete_worksheet": "删除工作表",
        "insert_rows": "插入行",
        "delete_rows": "删除行",
        "hide_rows": "隐藏行",
        "unhide_rows": "取消隐藏行",
        "delete_hidden_rows": "删除隐藏行",
        "insert_columns": "插入列",
        "delete_columns": "删除列",
        "hide_columns": "隐藏列",
        "unhide_columns": "取消隐藏列",
        "delete_hidden_columns": "删除隐藏列"
    }
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 步骤编号
    ws.column_dimensions['B'].width = 20  # 操作类型
    ws.column_dimensions['C'].width = 10  # 执行结果
    ws.column_dimensions['D'].width = 60  # 详细信息
    
    # 设置表头
    headers = ["步骤", "操作", "执行结果", "详细信息"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # 设置表头样式
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    # 填充数据
    for row, result in enumerate(step_results, 2):
        # 步骤编号
        ws.cell(row=row, column=1, value=result['step']).alignment = Alignment(horizontal="center")
        
        # 操作类型 - 显示完整的操作描述（包括位置信息等）
        operation_name = result['operation']
        operation_params = result.get('params', {})
        
        # 创建一个临时的StepItem对象来获取完整描述
        from models import StepItem
        temp_step = StepItem(operation_name, operation_params)
        operation_desc = str(temp_step)
        
        ws.cell(row=row, column=2, value=operation_desc)
        
        # 执行结果
        success_cell = ws.cell(row=row, column=3, value="成功" if result['success'] else "失败")
        success_cell.alignment = Alignment(horizontal="center")
        if result['success']:
            success_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            success_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 使用公共函数处理消息格式
        message = format_result_message(result)
        ws.cell(row=row, column=4, value=message)
    
    # 设置所有单元格的边框
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for row in range(1, len(step_results) + 2):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
    
    # 保存报告
    report_path = os.path.join(output_dir, file_name)
    wb.save(report_path)
    
    return report_path