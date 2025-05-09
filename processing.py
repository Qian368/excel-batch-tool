#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
处理线程模块
提供Excel处理的后台线程实现
"""

import traceback
import json
import os
from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl.utils.cell

from utils import parse_range_string


class ProcessingThread(QThread):
    """处理线程，用于在后台执行Excel处理操作"""
    progress_updated = pyqtSignal(int)
    operation_complete = pyqtSignal(bool, str)
    step_results_updated = pyqtSignal(list)  # 新增：步骤结果信号
    
    def __init__(self, processor, steps, file_paths):
        super().__init__()
        self.processor = processor
        self.steps = steps
        self.file_paths = file_paths
        # self.processor.output_dir = None  # 初始化输出目录属性
        self.step_results = []  # 新增：用于记录每个步骤的执行结果
    
    def run(self):
        try:
            self.step_results = []  # 每次运行前清空
            # 加载工作簿（会自动创建临时文件）
            self.processor.load_workbooks(self.file_paths)
            self.progress_updated.emit(20)

            # 执行每个步骤
            total_steps = len(self.steps)
            for i, step in enumerate(self.steps, 1):
                progress = 20 + (i / total_steps) * 60
                self.progress_updated.emit(int(progress))
                step_success = True
                step_msg = ""
                try:
                    # 处理步骤的实际操作名称和参数
                    operation_name = step.operation
                    params = step.params
                    
                    # 如果params中包含operation字段，则使用它作为实际操作名称
                    if isinstance(params, dict) and 'operation' in params:
                        operation_name = params.get('operation')
                        # 如果params中还包含params字段，则使用它作为实际参数
                        if 'params' in params:
                            params = params.get('params', {})
                    
                    # 根据操作类型执行相应的处理
                    if operation_name == 'convert_formulas_to_values':
                        self.processor.convert_formulas_to_values(self.file_paths)
                    
                    # 合并单元格处理操作
                    elif operation_name == 'merge_cells' or operation_name.startswith('合并单元格'):
                        # 合并单元格操作
                        # 如果操作名称以'合并单元格'开头，尝试从操作名称中提取范围
                        range_str = params.get('range_str', '')
                        if operation_name.startswith('合并单元格') and not range_str:
                            # 尝试从操作名称中提取范围，格式如：合并单元格(A1:B2)
                            import re
                            match = re.search(r'\(([A-Za-z0-9:]+)\)', operation_name)
                            if match:
                                range_str = match.group(1)
                        
                        sheet_indexes = params.get('sheet_indexes', [0])
                        
                        try:
                            # 检查core.py中是否有merge_cells方法，如果没有则实现一个基本版本
                            if hasattr(self.processor, 'merge_cells'):
                                self.processor.merge_cells(self.file_paths, sheet_indexes, range_str)
                            else:
                                # 如果没有merge_cells方法，则在这里实现基本功能
                                from openpyxl.utils import range_boundaries
                                
                                for file_path in self.file_paths:
                                    wb = self.processor.workbooks.get(file_path)
                                    if not wb:
                                        continue
                                        
                                    for sheet_index in sheet_indexes:
                                        if isinstance(sheet_index, int) and 0 <= sheet_index < len(wb.sheetnames):
                                            sheet = wb[wb.sheetnames[sheet_index]]
                                            
                                            # 获取合并范围的边界
                                            min_col, min_row, max_col, max_row = range_boundaries(range_str)
                                            
                                            # 保存左上角单元格的值
                                            top_left_value = sheet.cell(row=min_row, column=min_col).value
                                            
                                            # 合并单元格
                                            sheet.merge_cells(range_str)
                                            
                                            # 确保合并后的单元格保留左上角的值
                                            merged_cell = sheet.cell(row=min_row, column=min_col)
                                            merged_cell.value = top_left_value
                        except Exception as e:
                            error_msg = str(e)
                            if "Invalid cell coordinate" in error_msg:
                                raise ValueError(f"无效的单元格坐标: {range_str}")
                            elif "Cannot merge cells" in error_msg:
                                raise ValueError(f"无法合并单元格: {range_str}，可能与现有合并单元格冲突")
                            else:
                                raise
                    elif operation_name == 'process_merged_cells_all':
                        self.processor.process_merged_cells(
                            self.file_paths,
                            action=params.get('action', 'unmerge'),
                            mode='all'
                        )
                    elif operation_name == 'process_merged_cells_specific':
                        try:
                            range_str = params.get('range_str', '')
                            action = params.get('action', 'unmerge')
                            
                            # 检查单元格范围格式是否有效
                            if not range_str or not isinstance(range_str, str):
                                raise ValueError(f"无效的单元格范围格式: {range_str}")
                                
                            # 支持单个单元格输入，如B1
                            if ':' not in range_str:
                                # 尝试解析单个单元格坐标
                                try:
                                    # 使用openpyxl的函数解析单元格坐标
                                    cell_coord = openpyxl.utils.cell.coordinate_to_tuple(range_str)
                                    # 单个单元格也是有效的输入
                                except Exception:
                                    raise ValueError(f"无效的单元格坐标: {range_str}")
                            
                            # 标记是否有未合并单元格的情况
                            not_merged_files = []
                            
                            # 使用新的交集检测功能检查每个文件中的指定范围是否与合并单元格有交集
                            for file_path in self.file_paths:
                                wb = self.processor.workbooks.get(file_path)
                                if not wb:
                                    continue
                                    
                                for sheet_name in wb.sheetnames:
                                    sheet = wb[sheet_name]
                                    
                                    # 使用find_intersections方法检查是否有交集
                                    intersections = self.processor.find_intersections(sheet, range_str)
                                    
                                    # 如果没有交集，检查是否是完全匹配的合并单元格
                                    is_merged = False
                                    if not intersections:
                                        for merged_range in sheet.merged_cells.ranges:
                                            if str(merged_range) == range_str:
                                                is_merged = True
                                                break
                                    else:
                                        # 有交集，标记为已合并
                                        is_merged = True
                                        has_intersection = True
                                    
                                    if not is_merged:
                                        not_merged_files.append(file_path)
                                        break
                            
                            # 如果所有文件都没有合并单元格，则报告失败
                            # 但是，如果有交集的情况，应该继续执行拆分操作
                            has_intersection = False
                            
                            # 再次检查是否有交集的情况
                            for file_path in not_merged_files.copy():  # 使用副本进行迭代
                                wb = self.processor.workbooks.get(file_path)
                                if not wb:
                                    continue
                                    
                                for sheet_name in wb.sheetnames:
                                    sheet = wb[sheet_name]
                                    
                                    # 解析指定范围
                                    try:
                                        # 处理单个单元格的情况
                                        if ':' not in range_str:
                                            # 解析单个单元格坐标
                                            cell_coord = openpyxl.utils.cell.coordinate_to_tuple(range_str)
                                            min_row = max_row = cell_coord[0]
                                            min_col = max_col = cell_coord[1]
                                        else:
                                            # 处理范围的情况
                                            start_cell, end_cell = range_str.split(':')
                                            start_coord = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
                                            end_coord = openpyxl.utils.cell.coordinate_to_tuple(end_cell)
                                            
                                            min_row, min_col = start_coord[0], start_coord[1]
                                            max_row, max_col = end_coord[0], end_coord[1]
                                        
                                        # 检查是否有交集
                                        for merged_range in sheet.merged_cells.ranges:
                                            merged_min_row, merged_min_col = merged_range.min_row, merged_range.min_col
                                            merged_max_row, merged_max_col = merged_range.max_row, merged_range.max_col
                                            
                                            # 检查是否有交集
                                            if not (merged_max_row < min_row or merged_min_row > max_row or 
                                                    merged_max_col < min_col or merged_min_col > max_col):
                                                has_intersection = True
                                                not_merged_files.remove(file_path)  # 从未合并列表中移除
                                                break
                                    except Exception:
                                        pass
                                    
                                    if has_intersection:
                                        break
                                        
                                if has_intersection:
                                    break
                            
                            # 如果所有文件都没有合并单元格且没有交集，则报告失败
                            if len(not_merged_files) == len(self.file_paths) and not has_intersection:
                                step_success = False
                                step_msg = f"步骤{i}: {str(step)} 执行失败: 单元格范围 {range_str} 未合并，无法执行拆分操作"
                                self.step_results.append({
                                    'step': i,
                                    'operation': step.operation,
                                    'params': step.params,
                                    'success': step_success,
                                    'message': step_msg
                                })
                                continue
                            
                            # 执行拆分操作
                            self.processor.process_merged_cells(
                                self.file_paths,
                                action=action,
                                mode='specific',
                                range_str=range_str
                            )
                            
                            # 如果有部分文件没有合并单元格，添加警告信息
                            if not_merged_files:
                                file_names = [os.path.basename(f) for f in not_merged_files]
                                warning_msg = f"单元格范围 {range_str} 在以下文件中未合并: {', '.join(file_names)}"
                                step_msg = f"步骤{i}: {str(step)} 执行成功 - {warning_msg}"
                            else:
                                step_msg = f"步骤{i}: {str(step)} 执行成功"
                        except Exception as e:
                            # 检查是否是未合并单元格的错误
                            error_msg = str(e)
                            if "is not merged" in error_msg:
                                # 提取单元格范围信息
                                range_info = error_msg.split("Cell range ")[1].split(" is not")[0] if "Cell range " in error_msg else range_str
                                step_success = False
                                step_msg = f"步骤{i}: {str(step)} 执行失败: 单元格范围 {range_info} 未合并，无法执行拆分操作"
                            elif "Invalid cell coordinate" in error_msg:
                                step_success = False
                                step_msg = f"步骤{i}: {str(step)} 执行失败: 无效的单元格坐标: {range_str}"
                            elif "Cannot merge cells" in error_msg:
                                step_success = False
                                step_msg = f"步骤{i}: {str(step)} 执行失败: 无法处理单元格: {range_str}，可能与现有合并单元格冲突"
                            elif "'NoneType' object has no attribute" in error_msg:
                                step_success = False
                                step_msg = f"步骤{i}: {str(step)} 执行失败: 无法处理单元格: {range_str}，请检查工作表是否存在"
                            else:
                                # 其他错误直接抛出
                                step_success = False
                                step_msg = f"步骤{i}: {str(step)} 执行失败: {error_msg}"
                            
                            self.step_results.append({
                                'step': i,
                                'operation': step.operation,
                                'params': step.params,
                                'success': step_success,
                                'message': step_msg
                            })
                            continue
                    # 已在前面的条件分支中处理了merge_cells和以'合并单元格'开头的操作
                    
                    # 工作表操作
                    elif operation_name == 'create_worksheet' or operation_name.startswith('新建工作表'):
                        try:
                            sheet_name = params.get('sheet_name', '')
                            if not sheet_name:
                                raise ValueError("工作表名称不能为空")
                            self.processor.create_worksheet(self.file_paths, sheet_name)
                            step_msg = f"步骤{i}: 新建工作表（工作表：{sheet_name}） 执行成功"
                        except Exception as e:
                            step_success = False
                            error_msg = str(e)
                            if "already exists" in error_msg:
                                step_msg = f"步骤{i}: 新建工作表（工作表：{sheet_name}） 执行失败: 工作表已存在"
                            else:
                                step_msg = f"步骤{i}: 新建工作表（工作表：{sheet_name}） 执行失败: {error_msg}"
                            raise ValueError(step_msg)
                    elif operation_name == 'delete_worksheet' or operation_name.startswith('删除工作表'):
                        try:
                            sheet_name = params.get('sheet_name', '')
                            if not sheet_name:
                                raise ValueError("工作表名称不能为空")
                            
                            # 标记是否有工作表不存在的情况
                            sheet_not_exist_files = []
                            
                            # 检查core.py中是否有delete_worksheet方法，如果没有则实现一个基本版本
                            if hasattr(self.processor, 'delete_worksheet'):
                                # 在调用前记录工作表不存在的文件
                                for file_path in self.file_paths:
                                    wb = self.processor.workbooks.get(file_path)
                                    if not wb:
                                        continue
                                    if sheet_name not in wb.sheetnames:
                                        sheet_not_exist_files.append(file_path)
                                
                                # 如果所有文件都不存在该工作表，则抛出错误
                                if len(sheet_not_exist_files) == len(self.file_paths):
                                    raise ValueError(f"不存在该工作表：{sheet_name}")
                                
                                # 执行删除操作
                                self.processor.delete_worksheet(self.file_paths, sheet_name)
                            else:
                                # 如果没有delete_worksheet方法，则在这里实现基本功能
                                for file_path in self.file_paths:
                                    wb = self.processor.workbooks.get(file_path)
                                    if not wb:
                                        continue
                                        
                                    if sheet_name not in wb.sheetnames:
                                        sheet_not_exist_files.append(file_path)
                                        continue
                                        
                                    # 删除工作表
                                    del wb[sheet_name]
                            
                            # 根据结果设置消息
                            if sheet_not_exist_files:
                                # 有工作表不存在的情况，应当视为失败
                                file_names = [os.path.basename(f) for f in sheet_not_exist_files]
                                warning_msg = f"工作表 '{sheet_name}' 在以下文件中不存在: {', '.join(file_names)}"
                                step_success = False
                                step_msg = f"步骤{i}: 删除工作表（工作表：{sheet_name}） 执行失败 - {warning_msg}"
                            else:
                                step_msg = f"步骤{i}: 删除工作表（工作表：{sheet_name}） 执行成功"
                        except Exception as e:
                            step_success = False
                            error_msg = str(e)
                            if "does not exist" in error_msg or "不存在" in error_msg:
                                # 这种情况应该已经被上面的代码处理，不应该到这里
                                step_msg = f"步骤{i}: 删除工作表（工作表：{sheet_name}） 执行失败: 工作表不存在"
                            elif "cannot be deleted" in error_msg:
                                step_msg = f"步骤{i}: 删除工作表（工作表：{sheet_name}） 执行失败: 工作表无法删除，可能是工作簿中的唯一工作表"
                            else:
                                step_msg = f"步骤{i}: 删除工作表（工作表：{sheet_name}） 执行失败: {error_msg}"
                            raise ValueError(step_msg)
                    
                    # 行操作
                    elif operation_name in ['insert_rows', 'delete_rows', 'hide_rows', 'unhide_rows']:
                        positions = parse_range_string(params.get('position', '1'))
                        sheet_indexes = params.get('sheet_indexes', [0])
                        
                        # 获取合并单元格处理模式，仅在删除行时需要
                        merge_mode = None
                        if operation_name == 'delete_rows':
                            merge_mode = params.get('merge_mode', 'ignore')
                            # 确保merge_mode是有效的值
                            if merge_mode not in ['ignore', 'unmerge_only', 'unmerge_keep_value']:
                                merge_mode = 'ignore'
                        
                        for pos in positions:
                            if isinstance(pos, tuple):
                                start, end = pos
                                count = int(end) - int(start) + 1
                                # 调用相应的处理方法
                                if operation_name == 'delete_rows':
                                    self.processor.delete_rows(
                                        self.file_paths, sheet_indexes, int(start), count, merge_mode=merge_mode
                                    )
                                else:
                                    getattr(self.processor, operation_name)(
                                        self.file_paths, sheet_indexes, int(start), count
                                    )
                            else:
                                # 单个位置
                                if operation_name == 'delete_rows':
                                    self.processor.delete_rows(
                                        self.file_paths, sheet_indexes, int(pos), 1, merge_mode=merge_mode
                                    )
                                else:
                                    getattr(self.processor, operation_name)(
                                        self.file_paths, sheet_indexes, int(pos), 1
                                    )
                    
                    # 列操作
                    elif operation_name in ['insert_columns', 'delete_columns', 'hide_columns', 'unhide_columns']:
                        positions = parse_range_string(params.get('position', 'A'))
                        sheet_indexes = params.get('sheet_indexes', [0])
                        
                        # 获取合并单元格处理模式，仅在删除列时需要
                        merge_mode = None
                        if operation_name == 'delete_columns':
                            merge_mode = params.get('merge_mode', 'ignore')
                            # 确保merge_mode是有效的值
                            if merge_mode not in ['ignore', 'unmerge_only', 'unmerge_keep_value']:
                                merge_mode = 'ignore'
                        
                        for pos in positions:
                            if isinstance(pos, tuple):
                                start, end = pos
                                start_idx = ord(start.upper()) - ord('A') + 1
                                end_idx = ord(end.upper()) - ord('A') + 1
                                count = end_idx - start_idx + 1
                                
                                # 调用相应的处理方法
                                if operation_name == 'delete_columns':
                                    self.processor.delete_columns(
                                        self.file_paths, sheet_indexes, start_idx, count, merge_mode=merge_mode
                                    )
                                else:
                                    getattr(self.processor, operation_name)(
                                        self.file_paths, sheet_indexes, start_idx, count
                                    )
                            else:
                                # 单个位置
                                col_idx = ord(pos.upper()) - ord('A') + 1
                                
                                if operation_name == 'delete_columns':
                                    self.processor.delete_columns(
                                        self.file_paths, sheet_indexes, col_idx, 1, merge_mode=merge_mode
                                    )
                                else:
                                    getattr(self.processor, operation_name)(
                                        self.file_paths, sheet_indexes, col_idx, 1
                                    )
                    else:
                        # 未知操作类型
                        raise ValueError(f"未知的操作类型: {operation_name}")
                    step_msg = f"步骤{i}: {str(step)} 执行成功"
                except Exception as step_e:
                    step_success = False
                    error_msg = str(step_e)
                    
                    # 将英文错误信息转换为友好的中文提示
                    if "is not merged" in error_msg:
                        # 已在process_merged_cells_specific中处理
                        step_msg = f"步骤{i}: {str(step)} 执行失败: {error_msg}"
                    elif "does not exist" in error_msg and "Sheet" in error_msg:
                        # 工作表不存在错误
                        sheet_name = error_msg.split("Sheet ")[1].split(" does not")[0] if "Sheet " in error_msg else ""
                        step_msg = f"步骤{i}: {str(step)} 执行失败: 工作表 '{sheet_name}' 不存在"
                    elif "already exists" in error_msg and "Sheet" in error_msg:
                        # 工作表已存在错误
                        sheet_name = error_msg.split("Sheet ")[1].split(" already")[0] if "Sheet " in error_msg else ""
                        step_msg = f"步骤{i}: {str(step)} 执行失败: 工作表 '{sheet_name}' 已存在"
                    elif "Invalid cell coordinate" in error_msg:
                        # 无效的单元格坐标
                        step_msg = f"步骤{i}: {str(step)} 执行失败: 无效的单元格坐标"
                    else:
                        # 其他错误保持原样
                        step_msg = f"步骤{i}: {str(step)} 执行失败: {error_msg}"
                self.step_results.append({
                    'step': i,
                    'operation': step.operation,
                    'params': step.params,
                    'success': step_success,
                    'message': step_msg
                })

            # 保存工作簿
            try:
                self.processor.save_workbooks()
                self.progress_updated.emit(90)
                
                # 关闭工作簿
                self.processor.close_workbooks()
                self.progress_updated.emit(100)

                # 通过信号传递所有步骤结果
                self.step_results_updated.emit(self.step_results)

                # 发送操作完成信号
                self.operation_complete.emit(True, f"操作完成！文件已保存至输出目录。")

            except PermissionError as e:
                error_msg = f"文件保存权限错误: {str(e)}"
                print(error_msg)
                self.operation_complete.emit(False, error_msg)
                return
            except Exception as e:
                error_msg = f"文件保存过程中出错: {str(e)}\n{traceback.format_exc()}"
                print(error_msg)
                self.operation_complete.emit(False, error_msg)
                return

        except PermissionError as e:
            error_msg = f"处理过程中权限错误: {str(e)}"
            print(error_msg)
            self.operation_complete.emit(False, error_msg)
        except Exception as e:
            error_msg = f"处理过程中出错: {str(e)}\n{traceback.format_exc()}"
            print(error_msg)
            self.operation_complete.emit(False, error_msg)